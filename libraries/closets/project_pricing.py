import bpy
from bpy.types import Panel
from bpy.types import Operator
from bpy.props import (StringProperty,
                       BoolProperty,
                       IntProperty,
                       FloatProperty,
                       FloatVectorProperty,
                       BoolVectorProperty,
                       PointerProperty,
                       CollectionProperty,
                       EnumProperty)
import os, sys, subprocess
from snap import sn_types, sn_unit, sn_utils, sn_db, sn_paths, sn_xml, bl_info
from . import closet_props
from . import closet_utils
import xml.etree.ElementTree as ET
import openpyxl
import pandas
import numpy
import platform
import pathlib


PRICING_PROPERTY_NAMESPACE = "sn_project_pricing"

SKU_NUMBER = ''
PART_LABEL_ID = ''
PART_NAME = ''
QUANTITY = 0
LENGTH = 0
WIDTH = 0
THICKNESS = 0
DESCRIPTION = ''
ROOM_DRAWING_NUMBER = ''
PART_TYPE = ''

NAME = ''
TYPE = ''
VALUE = ''

# Counters for edgebanding field
S_COUNT = []
L_COUNT = []
EDGEBANDING = []

R_MATERIAL_PRICES = []
R_HARDWARE_PRICES = []
R_ACCESSORY_PRICES = []
R_UPGRADED_PANEL_PRICES = []
R_SPECIAL_ORDER_PRICES = []
R_GLASS_PRICES = []
R_LABOR_PRICES = []
R_MATERIAL_SQUARE_FOOTAGE = []
R_MATERIAL_LINEAR_FOOTAGE = []
R_PROJECT_TOTAL_HARDWARE = []
R_PROJECT_TOTAL_ACCESSORIES = []
R_PROJECT_TOTAL_MATERIAL = []
R_PROJECT_TOTAL_SQUARE_FOOTAGE = []
R_PROJECT_TOTAL_LINEAR_FOOTAGE = []
R_PROJECT_TOTAL_PRICE = []
R_ROOM_PRICING_LIST = []
R_PROJECT_TOTAL_UPGRADED_PANEL = []
R_PROJECT_TOTAL_LABOR = []
R_PROJECT_TOTAL_GLASS = []

F_MATERIAL_PRICES = []
F_HARDWARE_PRICES = []
F_ACCESSORY_PRICES = []
F_UPGRADED_PANEL_PRICES = []
F_SPECIAL_ORDER_PRICES = []
F_GLASS_PRICES = []
F_LABOR_PRICES = []
F_MATERIAL_SQUARE_FOOTAGE = []
F_MATERIAL_LINEAR_FOOTAGE = []
F_PROJECT_TOTAL_HARDWARE = []
F_PROJECT_TOTAL_ACCESSORIES = []
F_PROJECT_TOTAL_MATERIAL = []
F_PROJECT_TOTAL_SQUARE_FOOTAGE = []
F_PROJECT_TOTAL_LINEAR_FOOTAGE = []
F_PROJECT_TOTAL_PRICE = []
F_ROOM_PRICING_LIST = []
F_PROJECT_TOTAL_UPGRADED_PANEL = []
F_PROJECT_TOTAL_LABOR = []
F_PROJECT_TOTAL_GLASS = []

MATERIAL_PARTS_LIST = []
HARDWARE_PARTS_LIST = []
ACCESSORY_PARTS_LIST = []
UPGRADED_PANEL_PARTS_LIST = []
SPECIAL_ORDER_PARTS_LIST = []
GLASS_PARTS_LIST = []

material_types = ['PM', 'VN', 'EB', 'MD', 'WD', 'RE', 'GL', 'SN', 'PL', 'BB', 'QZ']
hardware_types = ['HW']
accessory_types = ['AC', 'WB', 'CM']
special_order_types = ['SO']
eb_orientation = ''


def set_job_info(root):
    # Job Information
    global JOB_NAME
    global JOB_NUMBER
    global CLIENT_NAME
    global CLIENT_ID
    global LEAD_ID
    global CLIENT_ADDRESS
    global CLIENT_CITY
    global CLIENT_STATE
    global CLIENT_ZIP
    global CLIENT_PHONE1
    global CLIENT_PHONE2
    global CLIENT_EMAIL
    global CLIENT_NOTES
    global CLIENT_DESIGNER
    global CLIENT_ROOM_COUNT
    global DESIGN_DATE

    name = root.find("Name")
    if name.text is not None:
        JOB_NAME = name.text

    for var in root.findall("Var"):
        var_iterator = iter(var)
        for _ in range(int(len(var) / 2)):
            NAME = next(var_iterator).text
            VALUE = next(var_iterator).text

            if NAME == 'jobnumber':
                JOB_NUMBER = VALUE
            if NAME == 'customername':
                CLIENT_NAME = VALUE
            if NAME == 'clientid':
                CLIENT_ID = VALUE
            if NAME == 'leadid':
                LEAD_ID = VALUE
            if NAME == 'projectaddress':
                CLIENT_ADDRESS = VALUE
            if NAME == 'city':
                CLIENT_CITY = VALUE
            if NAME == 'state':
                CLIENT_STATE = VALUE
            if NAME == 'zipcode':
                CLIENT_ZIP = VALUE
            if NAME == 'customerphone1':
                CLIENT_PHONE1 = VALUE
            if NAME == 'customerphone2':
                CLIENT_PHONE2 = VALUE
            if NAME == 'customeremail':
                CLIENT_EMAIL = VALUE
            if NAME == 'projectnotes':
                CLIENT_NOTES = VALUE
            if NAME == 'designer':
                CLIENT_DESIGNER = VALUE
            if NAME == 'totalroomcount':
                CLIENT_ROOM_COUNT = VALUE
            if NAME == 'designdate':
                DESIGN_DATE = VALUE


def get_project_xml(self):
    props = bpy.context.window_manager.sn_project
    proj = props.get_project()
    cleaned_name = proj.get_clean_name(proj.name)
    if proj.lead_id:
        LEAD_ID = proj.lead_id
    project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
    selected_project = os.path.join(project_dir, cleaned_name)
    xml_file = os.path.join(selected_project, "snap_job.xml")
    global PROJECT_NAME
    PROJECT_NAME = cleaned_name

    if not os.path.exists(project_dir):
        print("Projects Directory does not exist")
        # os.makedirs(project_dir)
    if not os.path.exists(xml_file):
        print("The 'snap_job.xml' file is not found. Please select desired rooms and perform an export.")
        return None
    else:
        return xml_file


def get_pricing_props():
    """ 
    This is a function to get access to all of the scene properties that are registered in this library
    """
    props = eval("bpy.context.scene." + PRICING_PROPERTY_NAMESPACE)
    return props


def get_square_footage(length_inches, width_inches):
    return round((width_inches * length_inches) / 144, 2)


def get_linear_footage(length_inches):
    return round((length_inches) / 12, 2)


def get_glass_inset_size(length, width, center_rail, style_name):
    group_1 = (
        "rome", "napoli", "carrara", "florence", "palermo", "venice",
        "colina", "portofino"
    )

    group_2 = (
        "verona", "sienna", "milano", "merano", "volterra", "bergamo",
        "aviano", "capri", "san marino", "molino vecchio", "pisa", "moderno"
    )

    group_3 = ("traviso",)

    # Case insensitive search
    style_name = style_name.replace(" Door Glass", "")
    name = style_name.lower()

    # Define adjustment values for different groups
    group_dimensions = {
        group_1: {'length': 5.125, 'width': 5.125},
        group_2: {'length': 3.63, 'width': 3.63},
        group_3: {'length': 3.88, 'width': 3.88}
    }

    # Check if the name belongs to any group
    for group, dimensions in group_dimensions.items():
        if name in group:
            adjustment = dimensions
            adjusted_length = float(length) - adjustment['length']
            adjusted_width = float(width) - adjustment['width']

            # Tall doors with center rail - Minus an additional 1.25" divided by 2
            if center_rail:
                center_rail_adjustment = 1.25
                adjusted_length = (adjusted_length - center_rail_adjustment) / 2

            print(
                style_name,
                "door glass calculation: door width minus",
                str(adjustment['width']) + '",',
                "door length minus",
                str(adjustment['length']) + '"',
                "minus an additional " + str(center_rail_adjustment) + '" and divided in 2 for center rail' if center_rail else "")

            adjusted_length =  round(adjusted_length, 2)
            adjusted_width = round(adjusted_width, 2)

            return adjusted_length, adjusted_width

    # Melamine door glass
    if center_rail:
        adjusted_length = round((float(length) - 6.75) / 2, 2)
    else:
        adjusted_length = float(length) - 4.75

    adjusted_width = round(float(width) - 4.75, 2)

    return adjusted_length, adjusted_width


def get_admin_fee():
    franchise_location = bpy.context.preferences.addons['snap'].preferences.franchise_location

    # Update Dallas admin fee while still using Phoenix pricing table
    if franchise_location == 'DAL':
        admin_fee = 0.20
    else:
        admin_fee = 0.15

    return admin_fee


def get_admin_fee_string():
    return "{:.0%}".format(get_admin_fee())


def price_check(sku_num, franchise, retail):
    if franchise > retail:
        print("$$$$$$$$$$$$$$$$ Price discrepancy within database $$$$$$$$$$$$$$$$")
        print("Sku Number: " + str(sku_num) + "  Franchise Price: " + str(franchise) + "  Retail Price: " + str(retail))


def set_column_width(sheet):
    # try:
    #     import openpyxl
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)
    ws = sheet
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                cell.font = openpyxl.styles.Font(bold=True)
            if cell.value or cell.value == 0 or cell.value == '0':
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 3
    # Code to freeze header row and first column
    ws.freeze_panes = 'A2'
    ws.freeze_panes = 'B2'


def protect_pricing(parts_file):
    wb = openpyxl.load_workbook(parts_file)

    if "Retail Pricing Summary" in wb.sheetnames:
        ws = wb["Retail Pricing Summary"]
        ws.protection.password = 'Cla$$y123'
        ws.protection.sheet = True
        ws1 = wb["ERP Pricing Summary"]
        ws2 = wb["Materials"]
        ws3 = wb["Hardware"]
        ws4 = wb["Accessories"]
        ws5 = wb["Upgraded Panels"]
        ws6 = wb["Glass"]
        ws7 = wb["Special Order"]
        ws1.sheet_state = 'hidden'
        ws2.sheet_state = 'hidden'
        ws3.sheet_state = 'hidden'
        ws4.sheet_state = 'hidden'
        ws5.sheet_state = 'hidden'
        ws6.sheet_state = 'hidden'

        # Add List of room names to Column N to create drop down list for special order parts and hide it
        for i in range(len(R_ROOM_PRICING_LIST)):
            ws["N" + str(i+1)] = R_ROOM_PRICING_LIST[i][0]
        ws.column_dimensions['N'].hidden = True

        data_val = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1="=OFFSET('Retail Pricing Summary'!$N$1,0,0,COUNTA('Retail Pricing Summary'!$N:$N),1)")
        ws7.add_data_validation(data_val)
        row_start = ws7.max_row
        for i in range(30):
            data_val.add(ws7["A" + str(row_start + (i+1))])
            ws7["M" + str(row_start + (i+1))].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws7["O" + str(row_start + (i+1))].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    if "Franchise Pricing Summary" in wb.sheetnames:
        ws = wb["Franchise Pricing Summary"]
        ws7 = wb["Special Order"]

        # Add List of room names to Column N to create drop down list for special order parts and hide it
        for i in range(len(F_ROOM_PRICING_LIST)):
            ws["N" + str(i+1)] = F_ROOM_PRICING_LIST[i][0]
        ws.column_dimensions['N'].hidden = True

        data_val = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1="=OFFSET('Franchise Pricing Summary'!$N$1,0,0,COUNTA('Franchise Pricing Summary'!$N:$N),1)")
        ws7.add_data_validation(data_val)
        row_start = ws7.max_row
        for i in range(30):
            data_val.add(ws7["A" + str(row_start + (i+1))])
            ws7["M" + str(row_start + (i+1))].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws7["O" + str(row_start + (i+1))].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    if not COS_FLAG:
        wb.security.workbookPassword = 'Cla$$y123'
        wb.security.lockStructure = True
            
    wb.save(filename=parts_file)

   
def display_parts_summary(parts_file):
    # Set currency formatting for Summary sheets
    wb = openpyxl.load_workbook(parts_file)
    row_start = 4
    with pandas.ExcelWriter(parts_file, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        if 'Materials Summary' in writer.sheets:
            materials = writer.sheets['Materials Summary']
            for i in range(materials.max_row):
                materials['H' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                materials['L' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                materials['N' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        if 'Drawers Summary' in writer.sheets:
            drawers = writer.sheets['Drawers Summary']
            if "Retail Pricing Summary" in wb.sheetnames:
                for i in range(drawers.max_row):
                    drawers['H' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                    drawers['L' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                    drawers['N' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            else:
                for i in range(drawers.max_row):
                    drawers['H' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                    drawers['L' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                    drawers['M' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                    drawers['O' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        if 'Hardware Summary' in writer.sheets:
            hardware = writer.sheets['Hardware Summary']
            for i in range(hardware.max_row):
                hardware['E' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                hardware['G' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        if 'Accessories Summary' in writer.sheets:
            accessories = writer.sheets['Accessories Summary']
            for i in range(accessories.max_row):
                accessories['F' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                accessories['H' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        if 'Upgraded Panels Summary' in writer.sheets:
            upgraded_panels = writer.sheets['Upgraded Panels Summary']
            for i in range(upgraded_panels.max_row):
                upgraded_panels['F' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                upgraded_panels['H' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                upgraded_panels['J' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        if 'Glass Summary' in writer.sheets:
            glass = writer.sheets['Glass Summary']
            for i in range(glass.max_row):
                glass['F' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
                glass['I' + str(row_start + i)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE  

    wb.save(filename=parts_file)
    # Password Protect pricing workbook and sheets
    protect_pricing(parts_file)
    
    if COS_FLAG and platform.system() == 'Windows':
        os.startfile(parts_file)
    if COS_FLAG:
        print(parts_file)
        print("Excel file output to COS Folder")
    else:
        os.startfile(parts_file)


def generate_retail_pricing_summary(parts_file):
    # try:
    #     import openpyxl
    #     import et_xmlfile
    #     import pandas
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)

    red_fill = openpyxl.styles.PatternFill(bgColor="FFCCCB")

    wb = openpyxl.Workbook()
    pricing_sheet = wb.active
    pricing_sheet.title = "Retail Pricing Summary"
    pricing_sheet.HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
    pricing_sheet.HeaderFooter.oddHeader.center.text = "Pricing Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
    pricing_sheet.HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)

    cell = pricing_sheet.cell(row=1, column=1)
    cell.value = "{} - {}".format(PROJECT_NAME, LEAD_ID)
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    pricing_sheet.merge_cells('B1:C1')
    cell = pricing_sheet.cell(row=1, column=2)
    cell.value = "SNaP Version: {}.{}.{}".format(bl_info['version'][0], bl_info['version'][1], bl_info['version'][2])
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell.fill = openpyxl.styles.PatternFill("solid", start_color="bcbcbc")
    cell2 = pricing_sheet.cell(row=1, column=4)
    cell2.value = "Designer: {}".format(CLIENT_DESIGNER) 
    cell2.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell2.fill = openpyxl.styles.PatternFill("solid", start_color="bcbcbc")

    erp_sheet = wb.create_sheet()
    erp_sheet.sheet_state = 'hidden'
    erp_sheet.title = "ERP Pricing Summary"
    erp_sheet.HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
    erp_sheet.HeaderFooter.oddHeader.center.text = "ERP Pricing Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
    erp_sheet.HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
    row_start = 2
    erp_row = 0
    cell = erp_sheet.cell(row=1, column=1)
    cell.value = "{} - {}".format(PROJECT_NAME, LEAD_ID)
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    erp_sheet["A" + str(erp_row + 1)] = "Room Name"
    erp_sheet["A" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)
    erp_sheet["B" + str(erp_row + 1)] = "Suggested Retail"
    erp_sheet["B" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)
    erp_sheet["C" + str(erp_row + 1)] = "Quoted Customer Price"
    erp_sheet["C" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)
    erp_sheet["D" + str(erp_row + 1)] = "Franchise Price"
    erp_sheet["D" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)
    erp_sheet["E" + str(erp_row + 1)] = "Est Materials Cost"
    erp_sheet["E" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)
    erp_sheet["F" + str(erp_row + 1)] = "Room Drawing Number"
    erp_sheet["F" + str(erp_row + 1)].font = openpyxl.styles.Font(bold=True)

    for i in range(len(R_ROOM_PRICING_LIST)):
        pricing_sheet["A" + str(row_start + 1)] = "Room Name"
        pricing_sheet["A" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 1)] = R_ROOM_PRICING_LIST[i][0]
        pricing_sheet["C" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["I" + str(row_start)] = "Promotion Description"
        pricing_sheet["I" + str(row_start)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["I" + str(row_start)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start)] = ""
        pricing_sheet["J" + str(row_start)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["J" + str(row_start)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["I" + str(row_start + 1)] = "Adjustment %"
        pricing_sheet["I" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["J" + str(row_start + 1)] = "Adjustment Value"
        pricing_sheet["J" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 2)] = "Special Order Items Count"
        pricing_sheet["A" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 2)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 2)] = "=COUNTIF('Special Order'!A:A," + "C" + str(row_start + 1) + ")"
        pricing_sheet["A" + str(row_start + 3)] = "Materials Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 3)] = R_ROOM_PRICING_LIST[i][3]
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["I" + str(row_start + 3)] = 0
        pricing_sheet["I" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 3)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 3)] = "=I" + str(row_start + 3) + "*C" + str(row_start + 3)
        pricing_sheet["J" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["J" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 4)] = "Hardware Price"
        pricing_sheet["A" + str(row_start + 4)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 4)] = R_ROOM_PRICING_LIST[i][4]
        pricing_sheet["C" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["I" + str(row_start + 4)] = 0
        pricing_sheet["I" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 4)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 4)] = "=I" + str(row_start + 4) + "*C" + str(row_start + 4)
        pricing_sheet["J" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["J" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 5)] = "Accessories Price"
        pricing_sheet["A" + str(row_start + 5)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 5)] = R_ROOM_PRICING_LIST[i][5]
        pricing_sheet["C" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["I" + str(row_start + 5)] = 0
        pricing_sheet["I" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 5)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 5)] = "=I" + str(row_start + 5) + "*C" + str(row_start + 5)
        pricing_sheet["J" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["J" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 6)] = "Upgraded Panels Price"
        pricing_sheet["A" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 6)] = R_ROOM_PRICING_LIST[i][6]
        pricing_sheet["C" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["I" + str(row_start + 6)] = 0
        pricing_sheet["I" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 6)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 6)] = "=I" + str(row_start + 6) + "*C" + str(row_start + 6)
        pricing_sheet["J" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["J" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 7)] = "Glass Price"
        pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 7)] = R_ROOM_PRICING_LIST[i][9]
        pricing_sheet["C" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["I" + str(row_start + 7)] = 0
        pricing_sheet["I" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 7)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 7)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 7)] = "=I" + str(row_start + 7) + "*C" + str(row_start + 7)
        pricing_sheet["J" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["J" + str(row_start + 7)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 8)] = "Special Order Price"
        pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 8)] = "=SUMIF('Special Order'!A:A," + "C" + str(row_start + 1) + ",'Special Order'!O:O)"
        pricing_sheet["C" + str(row_start + 8)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["A" + str(row_start + 9)] = "Adjustments"
        pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 9)] = "=SUM(J" + str(row_start + 3) + ":J" + str(row_start + 7) + ")"
        pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["C" + str(row_start + 9)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 10)] = "Room Subtotal"
        pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 10)] = "=SUM(C" + str(row_start + 3) + ":C" + str(row_start + 8) + ")" + "-" + "C" + str(row_start + 9)
        pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["A" + str(row_start + 11)] = f"Admin Fee ({get_admin_fee_string()})"
        pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 11)] = f"=C{str(row_start + 10)}*{str(get_admin_fee())}"
        pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["A" + str(row_start + 12)] = "Adjusted Room Subtotal"
        pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 12)] = "=C" + str(row_start + 10) + "+C" + str(row_start + 11)
        pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["A" + str(row_start + 13)] = "O.D.C Price"
        pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 13)] = 0
        pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["C" + str(row_start + 13)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["A" + str(row_start + 14)] = "Tear Out Price"
        pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 14)] = 0
        pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
        pricing_sheet["C" + str(row_start + 14)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["A" + str(row_start + 15)] = "Room Grand Total"
        pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 15)] = "=" + "C" + str(row_start + 12) + "+C" + str(row_start + 13) + "+C" + str(row_start + 14)
        pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD

        erp_sheet["A" + str(erp_row + 2)] = R_ROOM_PRICING_LIST[i][0]
        erp_sheet["A" + str(erp_row + 2)].font = openpyxl.styles.Font(bold=True)
        erp_sheet["B" + str(erp_row + 2)] = "='Retail Pricing Summary'!" + "C" + str(row_start + 10)
        erp_sheet["B" + str(erp_row + 2)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        erp_sheet["C" + str(erp_row + 2)] = "='Retail Pricing Summary'!" + "C" + str(row_start + 15)
        erp_sheet["C" + str(erp_row + 2)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # erp_sheet["D" + str(erp_row + 2)] = "=C" + str(erp_row + 2) + "*0.3"
        erp_sheet["D" + str(erp_row + 2)] = F_ROOM_PRICING_LIST[i][3] + F_ROOM_PRICING_LIST[i][4] + F_ROOM_PRICING_LIST[i][5] + F_ROOM_PRICING_LIST[i][6] + F_ROOM_PRICING_LIST[i][9]
        erp_sheet["D" + str(erp_row + 2)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        erp_sheet["E" + str(erp_row + 2)] = "=D" + str(erp_row + 2) + "*0.5"
        erp_sheet["E" + str(erp_row + 2)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        erp_sheet["F" + str(erp_row + 2)] = R_ROOM_PRICING_LIST[i][11]

        pricing_sheet["A" + str(row_start + 17)] = ""
        row_start = pricing_sheet.max_row
        erp_row = erp_sheet.max_row

    pricing_sheet["B" + str(row_start + 2)] = "Unassigned Special Order Items"
    pricing_sheet["B" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 3)] = "Unassigned Special Order Price"
    pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["B" + str(row_start + 3)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 3)] = "=SUMIF('Special Order'!A:A," + '""' + ",'Special Order'!O:O)"
    pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet.conditional_formatting.add("C" + str(row_start + 3), openpyxl.formatting.rule.FormulaRule(formula=["C" + str(row_start + 3) + "<>0"], stopIfTrue=True, fill=red_fill))
    pricing_sheet["D" + str(row_start + 3)] = "=IF(" + "C" + str(row_start + 3) + "," + '"Please assign Room Names for these items"' + "," + '""' + ")"
    pricing_sheet.conditional_formatting.add("D" + str(row_start + 3), openpyxl.formatting.rule.FormulaRule(formula=["C" + str(row_start + 3) + "<>0"], stopIfTrue=True, fill=red_fill))

    pricing_sheet["B" + str(row_start + 6)] = "Project Totals"
    pricing_sheet["B" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 7)] = "Room Count Total"
    pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 7)] = len(R_ROOM_PRICING_LIST)
    pricing_sheet["A" + str(row_start + 8)] = "Special Order Items Count Total"
    pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
    # pricing_sheet["B" + str(row_start + 8)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 8)] = "=SUMIF('Special Order'!H:H," + '"' + ">0" + '"' + ")"
    pricing_sheet["A" + str(row_start + 9)] = "Materials Price Total"
    pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 9)] = "=SUMIF('Materials'!P:P," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 10)] = "Hardware Price Total"
    pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 10)] = "=SUMIF('Hardware'!K:K," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 11)] = "Accessories Price Total"
    pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 11)] = "=SUMIF('Accessories'!K:K," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 12)] = "Upgraded Panels Price Total"
    pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 12)] = "=SUMIF('Upgraded Panels'!O:O," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 13)] = "Glass Price Total"
    pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 13)] = "=SUMIF('Glass'!O:O," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 14)] = "Special Order Price Total"
    pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 14)] = "=SUMIF('Special Order'!O:O," + '"' + "<>0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 15)] = "Adjustments Total"
    pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 15)] = "=SUM(J:J)"
    pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["C" + str(row_start + 15)].font = openpyxl.styles.Font(color="00339966")
    pricing_sheet["A" + str(row_start + 16)] = "Project Subtotal"
    pricing_sheet["A" + str(row_start + 16)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 16)] = "=C" + str(row_start + 9) + "+C" + str(row_start + 10) + "+C" + str(row_start + 11) + "+C" + str(row_start + 12) + "+C" + str(row_start + 13) + "+C" + str(row_start + 14) +"-C" + str(row_start + 15)
    pricing_sheet["C" + str(row_start + 16)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 17)] = f"Admin Fee ({get_admin_fee_string()})"
    pricing_sheet["A" + str(row_start + 17)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 17)] = f"=C{str(row_start + 16)}*{str(get_admin_fee())}"
    pricing_sheet["C" + str(row_start + 17)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 18)] = "Adjusted Project Subtotal"
    pricing_sheet["A" + str(row_start + 18)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 18)] = "=C" + str(row_start + 16) + "+C" + str(row_start + 17)
    pricing_sheet["C" + str(row_start + 18)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 19)] = "O.D.C Price Total"
    pricing_sheet["A" + str(row_start + 19)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 19)] = "=SUMIF(A:A," + '"O.D.C Price"' + ",C:C)"
    pricing_sheet["C" + str(row_start + 19)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["C" + str(row_start + 19)].protection = openpyxl.styles.Protection(locked=False)
    pricing_sheet["A" + str(row_start + 20)] = "Tear Out Price Total"
    pricing_sheet["A" + str(row_start + 20)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 20)] = "=SUMIF(A:A," + '"Tear Out Price"' + ",C:C)"
    pricing_sheet["C" + str(row_start + 20)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["C" + str(row_start + 20)].protection = openpyxl.styles.Protection(locked=False)
    pricing_sheet["A" + str(row_start + 21)] = "Grand Total"
    pricing_sheet["A" + str(row_start + 21)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 21)] = "=C" + str(row_start + 18) + "+C" + str(row_start + 19) + "+C" + str(row_start + 20)
    pricing_sheet["C" + str(row_start + 21)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD

    set_column_width(pricing_sheet)
    set_column_width(erp_sheet)

    try:
        wb.save(filename=parts_file)
    except PermissionError:
        print("Retail Pricing Spreadsheet Open...Please close file: " + str(parts_file))
    print("Retail Pricing Summary Created")


def generate_franchise_pricing_summary(parts_file):
    # try:
    #     import openpyxl
    #     import et_xmlfile
    #     import pandas
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)

    red_fill = openpyxl.styles.PatternFill(bgColor="FFCCCB")

    wb = openpyxl.Workbook()
    pricing_sheet = wb.active
    pricing_sheet.title = "Franchise Pricing Summary"
    pricing_sheet.HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
    pricing_sheet.HeaderFooter.oddHeader.center.text = "Pricing Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
    pricing_sheet.HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)

    cell = pricing_sheet.cell(row=1, column=1)
    cell.value = "{} - {}".format(PROJECT_NAME, LEAD_ID)
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    pricing_sheet.merge_cells('B1:C1')
    cell = pricing_sheet.cell(row=1, column=2)
    cell.value = "SNaP Version: {}.{}.{}".format(bl_info['version'][0], bl_info['version'][1], bl_info['version'][2])
    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell.fill = openpyxl.styles.PatternFill("solid", start_color="bcbcbc")
    cell2 = pricing_sheet.cell(row=1, column=4)
    cell2.value = "Designer: {}".format(CLIENT_DESIGNER) 
    cell2.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cell2.fill = openpyxl.styles.PatternFill("solid", start_color="bcbcbc")

    row_start = 2

    for i in range(len(F_ROOM_PRICING_LIST)):
        pricing_sheet["A" + str(row_start + 1)] = "Room Name"
        pricing_sheet["A" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 1)] = F_ROOM_PRICING_LIST[i][0]
        pricing_sheet["C" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["H" + str(row_start)] = "Promotion Description"
        pricing_sheet["H" + str(row_start)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["H" + str(row_start)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["I" + str(row_start)] = ""
        pricing_sheet["I" + str(row_start)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["I" + str(row_start)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["H" + str(row_start + 1)] = "Adjustment %"
        pricing_sheet["H" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["I" + str(row_start + 1)] = "Adjustment Value"
        pricing_sheet["I" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 2)] = "Special Order Items Count"
        pricing_sheet["A" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 2)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 2)] = "=COUNTIF('Special Order'!A:A," + "C" + str(row_start + 1) + ")"
        pricing_sheet["A" + str(row_start + 3)] = "Materials Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 3)] = F_ROOM_PRICING_LIST[i][3]
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["I" + str(row_start + 3)] = 0
        pricing_sheet["I" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 3)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 3)] = "=I" + str(row_start + 3) + "*C" + str(row_start + 3)
        pricing_sheet["J" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["J" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 4)] = "Hardware Price"
        pricing_sheet["A" + str(row_start + 4)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 4)] = F_ROOM_PRICING_LIST[i][4]
        pricing_sheet["C" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["I" + str(row_start + 4)] = 0
        pricing_sheet["I" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 4)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 4)] = "=I" + str(row_start + 4) + "*C" + str(row_start + 4)
        pricing_sheet["J" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["J" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 5)] = "Accessories Price"
        pricing_sheet["A" + str(row_start + 5)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 5)] = F_ROOM_PRICING_LIST[i][5]
        pricing_sheet["C" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["I" + str(row_start + 5)] = 0
        pricing_sheet["I" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 5)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 5)] = "=I" + str(row_start + 5) + "*C" + str(row_start + 5)
        pricing_sheet["J" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["J" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 6)] = "Upgraded Panels Price"
        pricing_sheet["A" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 6)] = F_ROOM_PRICING_LIST[i][6]
        pricing_sheet["C" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["I" + str(row_start + 6)] = 0
        pricing_sheet["I" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 6)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 6)] = "=I" + str(row_start + 6) + "*C" + str(row_start + 6)
        pricing_sheet["J" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["J" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 7)] = "Glass Price"
        pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 7)] = F_ROOM_PRICING_LIST[i][9]
        pricing_sheet["C" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["I" + str(row_start + 7)] = 0
        pricing_sheet["I" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["I" + str(row_start + 7)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["I" + str(row_start + 7)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["J" + str(row_start + 7)] = "=I" + str(row_start + 7) + "*C" + str(row_start + 7)
        pricing_sheet["J" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["J" + str(row_start + 7)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 8)] = "Special Order Price"
        pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 8)] = "=SUMIF('Special Order'!A:A," + "C" + str(row_start + 1) +  ",'Special Order'!O:O)"
        pricing_sheet["C" + str(row_start + 8)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["A" + str(row_start + 9)] = "Adjustments"
        pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 9)] = "=SUM(J" + str(row_start + 3) + ":J" + str(row_start + 7) + ")"
        pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["C" + str(row_start + 9)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 10)] = "Room Subtotal"
        pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 10)] = "=SUM(C" + str(row_start + 3) + ":C" + str(row_start + 8) + ")" + "-" + "C" + str(row_start + 9)
        pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["A" + str(row_start + 11)] = "Admin Fee (0%)"
        pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 11)] = "=" + "C" + str(row_start + 10) + "*.0"
        pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["A" + str(row_start + 12)] = "Adjusted Room Subtotal"
        pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 12)] = "=" + "C" + str(row_start + 10) + "+C" + str(row_start + 11)
        pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["A" + str(row_start + 13)] = "O.D.C Price"
        pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 13)] = 0
        pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["C" + str(row_start + 13)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["A" + str(row_start + 14)] = "Tear Out Price"
        pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 14)] = 0
        pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["C" + str(row_start + 14)].protection = openpyxl.styles.Protection(locked=False)
        pricing_sheet["A" + str(row_start + 15)] = "Room Grand Total"
        pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 15)] = "=" + "C" + str(row_start + 12) + "+C" + str(row_start + 13) + "+C" + str(row_start + 14)
        pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD

        pricing_sheet["A" + str(row_start + 17)] = ""
        row_start = pricing_sheet.max_row

    pricing_sheet["B" + str(row_start + 2)] = "Unassigned Special Order Items"
    pricing_sheet["B" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 3)] = "Unassigned Special Order Price"
    pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["B" + str(row_start + 3)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 3)] = "=SUMIF('Special Order'!A:A," + '""' + ",'Special Order'!O:O)"
    pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet.conditional_formatting.add("C" + str(row_start + 3), openpyxl.formatting.rule.FormulaRule(formula=["C" + str(row_start + 3) + "<>0"], stopIfTrue=True, fill=red_fill))
    pricing_sheet["D" + str(row_start + 3)] = "=IF(" + "C" + str(row_start + 3) + "," + '"Please assign Room Names for these items"' + "," + '""' + ")"
    pricing_sheet.conditional_formatting.add("D" + str(row_start + 3), openpyxl.formatting.rule.FormulaRule(formula=["C" + str(row_start + 3) + "<>0"], stopIfTrue=True, fill=red_fill))

    pricing_sheet["B" + str(row_start + 6)] = "Project Totals"
    pricing_sheet["B" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 7)] = "Room Count Total"
    pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 7)] = len(F_ROOM_PRICING_LIST)
    pricing_sheet["A" + str(row_start + 8)] = "Special Order Items Count Total"
    pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
    # pricing_sheet["B" + str(row_start + 8)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 8)] = "=SUMIF('Special Order'!H:H," + '"' + ">0" + '"' + ")"
    pricing_sheet["A" + str(row_start + 9)] = "Materials Price Total"
    pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 9)] = "=SUMIF('Materials'!P:P," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 10)] = "Hardware Price Total"
    pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 10)] = "=SUMIF('Hardware'!K:K," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 11)] = "Accessories Price Total"
    pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 11)] = "=SUMIF('Accessories'!K:K," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 12)] = "Upgraded Panels Price Total"
    pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 12)] = "=SUMIF('Upgraded Panels'!O:O," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 13)] = "Glass Price Total"
    pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 13)] = "=SUMIF('Glass'!O:O," + '"' + ">0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 14)] = "Special Order Price Total"
    pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 14)] = "=SUMIF('Special Order'!O:O," + '"' + "<>0" + '"' + ")"
    pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 15)] = "Adjustments Total"
    pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 15)] = "=SUM(J:J)"
    pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["C" + str(row_start + 15)].font = openpyxl.styles.Font(color="00339966")
    pricing_sheet["A" + str(row_start + 16)] = "Project Subtotal"
    pricing_sheet["A" + str(row_start + 16)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 16)] = "=C" + str(row_start + 9) + "+C" + str(row_start + 10) + "+C" + str(row_start + 11) + "+C" + str(row_start + 12) + "+C" + str(row_start + 13) + "+C" + str(row_start + 14) +"-C" + str(row_start + 15)
    pricing_sheet["C" + str(row_start + 16)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD
    pricing_sheet["A" + str(row_start + 17)] = "Admin Fee (0%)"
    pricing_sheet["A" + str(row_start + 17)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 17)] = "=" + "C" + str(row_start + 16) + "*.0"
    pricing_sheet["C" + str(row_start + 17)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 18)] = "Adjusted Project Subtotal"
    pricing_sheet["A" + str(row_start + 18)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 18)] = "=" + "C" + str(row_start + 16) + "+C" + str(row_start + 17)
    pricing_sheet["C" + str(row_start + 18)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 19)] = "O.D.C Price Total"
    pricing_sheet["A" + str(row_start + 19)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 19)] = "=SUMIF(A:A," + '"O.D.C Price"' + ",C:C)"
    pricing_sheet["C" + str(row_start + 19)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["C" + str(row_start + 19)].protection = openpyxl.styles.Protection(locked=False)
    pricing_sheet["A" + str(row_start + 20)] = "Tear Out Price Total"
    pricing_sheet["A" + str(row_start + 20)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 20)] = "=SUMIF(A:A," + '"Tear Out Price"' + ",C:C)"
    pricing_sheet["C" + str(row_start + 20)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["C" + str(row_start + 20)].protection = openpyxl.styles.Protection(locked=False)
    pricing_sheet["A" + str(row_start + 21)] = "Grand Total"
    pricing_sheet["A" + str(row_start + 21)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 21)] = "=C" + str(row_start + 18) + "+C" + str(row_start + 19) + "+C" + str(row_start + 20)
    pricing_sheet["C" + str(row_start + 21)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD

    set_column_width(pricing_sheet)
    try:
        wb.save(filename=parts_file)
    except PermissionError:
        print("Pricing Spreadsheet Open...Please close file: " + str(parts_file))
    print("Franchise Pricing Summary Created")


def generate_parts_summary(parts_file, materials_sheet, hardware_sheet, accessories_sheet, upgraded_panel_sheet, glass_sheet, so_sheet):
    # try:
    #     import pandas
    #     import numpy
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)

    with pandas.ExcelWriter(parts_file, mode='a') as writer:
        if materials_sheet is not None:
            df_materials = pandas.read_excel(parts_file, sheet_name='Materials').query('"GL" <= SKU_NUMBER <= "WD~" \
                                                                                        and SKU_NUMBER.str.contains("BB")==False \
                                                                                        and PART_NAME.str.contains("DrwrBox")==False \
                                                                                        and (PART_NAME.str.contains("File Rail")==False and SKU_NUMBER.str.contains("EB")==False) \
                                                                                        and (PART_NAME.str.contains("Drawer Side")==False and SKU_NUMBER.str.contains("EB")==False) \
                                                                                        and (PART_NAME.str.contains("Drawer Sub Front")==False and SKU_NUMBER.str.contains("EB")==False) \
                                                                                        and (PART_NAME.str.contains("Drawer Back")==False and SKU_NUMBER.str.contains("EB")==False) \
                                                                                        and (PART_NAME.str.contains("Drawer Bottom")==False)',
                                                                                        engine='python')
            if df_materials.shape[0] > 0:
                materials_summary = pandas.pivot_table(
                    df_materials,
                    index=['ROOM_NAME', 'WALL_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS', 'EDGEBANDING', 'PART_PRICE', 'SQUARE_FT', 'DRILLING', 'PULL_DRILLING'],
                    values=['LABOR', 'QUANTITY', 'TOTAL_PRICE'],
                    aggfunc=[numpy.sum])

                if not materials_summary.empty:
                    materials_summary.to_excel(writer, sheet_name='Materials Summary')
                    writer.sheets['Materials Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                    writer.sheets['Materials Summary'].HeaderFooter.oddHeader.center.text = "Materials Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                    writer.sheets['Materials Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                    cell = writer.sheets['Materials Summary'].cell(row=1, column=1)
                    cell.value = "{} - {}".format(PROJECT_NAME, LEAD_ID)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    set_column_width(writer.sheets['Materials Summary'])
                else:
                    print("Drawers summary DataFrame is empty. Skipping writing to Excel.")

            df_drawers = pandas.read_excel(parts_file, sheet_name='Materials').query('(PART_NAME.str.contains("File Rail")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("Drawer Side")==True and SKU_NUMBER.str.contains("BB")==True) \
                                                                                    or (PART_NAME.str.contains("Drawer Sub Front")==True and SKU_NUMBER.str.contains("BB")==True) \
                                                                                    or (PART_NAME.str.contains("Drawer Back")==True and SKU_NUMBER.str.contains("BB")==True) \
                                                                                    or (PART_NAME.str.contains("Drawer Side")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("Drawer Sub Front")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("Drawer Back")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("Drawer Bottom")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Bottom - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Side - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Back - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Front - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Bttm DT - BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Side DT - BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Back DT - BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Front DT - BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Inset Bttm - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Side Inset Bttm- Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Back Inset Bttm- Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Front Inset Bttm - Mel")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Inset Bttm - BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Side Inset Bttm- BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Back Inset Bttm- BB")==True and SKU_NUMBER.str.contains("EB")==False) \
                                                                                    or (PART_NAME.str.contains("DrwrBox Front Inset Bttm - BB")==True and SKU_NUMBER.str.contains("EB")==False) ',
                                                                                    engine='python')

            if df_drawers.shape[0] > 0:
                wb = openpyxl.load_workbook(parts_file)
                if "Retail Pricing Summary" in wb.sheetnames:
                    drawers_summary = pandas.pivot_table(
                        df_drawers,
                        index=['ROOM_NAME', 'WALL_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS', 'EDGEBANDING', 'PART_PRICE', 'SQUARE_FT', 'LINEAR_FT', 'PULL_DRILLING'],
                        values=['LABOR', 'QUANTITY', 'TOTAL_PRICE'],
                        aggfunc=[numpy.sum])

                    if not drawers_summary.empty:
                        drawers_summary.to_excel(writer, sheet_name='Drawers Summary')
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.center.text = "Drawers Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                        set_column_width(writer.sheets['Drawers Summary'])
                    else:
                        print("Drawers summary DataFrame is empty. Skipping writing to Excel.")
                else:
                    print("Generating Franchise Drawers Summary")
                    drawers_summary = pandas.pivot_table(
                        df_drawers,
                        index=['ROOM_NAME', 'WALL_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS', 'EDGEBANDING', 'PART_PRICE', 'SQUARE_FT', 'LINEAR_FT', 'PULL_DRILLING', 'ASSEMBLED'],
                        values=['LABOR', 'QUANTITY', 'TOTAL_PRICE'],
                        aggfunc=[numpy.sum])

                    if not drawers_summary.empty:
                        drawers_summary.to_excel(writer, sheet_name='Drawers Summary')
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.center.text = "Drawers Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                        writer.sheets['Drawers Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                        set_column_width(writer.sheets['Drawers Summary'])
                    else:
                        print("Drawers summary DataFrame is empty. Skipping writing to Excel.")

        if hardware_sheet is not None:
            df_hardware = pandas.read_excel(parts_file, sheet_name='Hardware')
            if df_hardware.shape[0] > 0:
                hardware_summary = pandas.pivot_table(df_hardware, 
                                                        index=['ROOM_NAME', 'VENDOR_NAME', 'VENDOR_ITEM', 'PART_NAME', 'PART_PRICE'], 
                                                        values=['QUANTITY', 'TOTAL_PRICE'],
                                                        aggfunc=[numpy.sum])
                hardware_summary.to_excel(writer, sheet_name='Hardware Summary')
                writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.center.text = "Hardware Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                set_column_width(writer.sheets['Hardware Summary'])
            
        if accessories_sheet is not None:
            df_accessories = pandas.read_excel(parts_file, sheet_name='Accessories')
            if df_accessories.shape[0] > 0:
                accessories_summary = pandas.pivot_table(df_accessories, 
                                                            index=['ROOM_NAME', 'VENDOR_NAME', 'VENDOR_ITEM', 'PART_NAME', 'PART_DIMENSIONS', 'PART_PRICE'], 
                                                            values=['QUANTITY', 'TOTAL_PRICE'],
                                                            aggfunc=[numpy.sum])
                accessories_summary.to_excel(writer, sheet_name='Accessories Summary')
                writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.center.text = "Accessories Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                set_column_width(writer.sheets['Accessories Summary'])

        if upgraded_panel_sheet is not None:
            df_upgraded_panel = pandas.read_excel(parts_file, sheet_name='Upgraded Panels')
            if df_upgraded_panel.shape[0] > 0:
                upgraded_panel_summary = pandas.pivot_table(df_upgraded_panel, 
                                                            index=['ROOM_NAME', 'PAINT_STAIN_COLOR', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS', 'PART_PRICE', 'SQUARE_FT'], 
                                                            values=['LABOR', 'QUANTITY', 'TOTAL_PRICE'],
                                                            aggfunc=[numpy.sum])
                upgraded_panel_summary.to_excel(writer, sheet_name='Upgraded Panels Summary')
                writer.sheets['Upgraded Panels Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                writer.sheets['Upgraded Panels Summary'].HeaderFooter.oddHeader.center.text = "Upgraded Panels Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                writer.sheets['Upgraded Panels Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                set_column_width(writer.sheets['Upgraded Panels Summary'])

        if glass_sheet is not None:
            df_glass = pandas.read_excel(parts_file, sheet_name='Glass')
            if df_glass.shape[0] > 0:
                glass_summary = pandas.pivot_table(df_glass, 
                                                    index=['ROOM_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS', 'PART_PRICE', 'SQUARE_FT'], 
                                                    values=['QUANTITY', 'TOTAL_PRICE'],
                                                    aggfunc=[numpy.sum])
                glass_summary.to_excel(writer, sheet_name='Glass Summary')
                writer.sheets['Glass Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nLead ID: {}".format(CLIENT_NAME, LEAD_ID)
                writer.sheets['Glass Summary'].HeaderFooter.oddHeader.center.text = "Glass Summary Sheet\nDesigner Name: {}".format(CLIENT_DESIGNER)
                writer.sheets['Glass Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
                set_column_width(writer.sheets['Glass Summary'])

        # if so_sheet is not None:
        #     df_so = pandas.read_excel(parts_file, sheet_name='Special Order')
        #     if df_so.shape[0] > 0:
        #         so_summary = pandas.pivot_table(df_so, 
        #                                         index=['ROOM_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS'], 
        #                                         values=['QUANTITY'], 
        #                                         aggfunc=[numpy.sum])
        #         so_summary.to_excel(writer, sheet_name='Special Order Summary')
        #         writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
        #         writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.center.text = "Special Order Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
        #         writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
        #         set_column_width(writer.sheets['Special Order Summary'])

    print("Parts Summary Created")
    display_parts_summary(parts_file)


def generate_retail_parts_list():
    if not COS_FLAG:
        props = bpy.context.window_manager.sn_project
        proj = props.get_project()
        cleaned_name = proj.get_clean_name(proj.name)
        project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
        selected_project = os.path.join(project_dir, cleaned_name)
        parts_file = os.path.join(selected_project, "Retail_Pricing_Parts_List" + "(" + str(cleaned_name) + ").xlsx")

        if not os.path.exists(project_dir):
            print("Projects Directory does not exist")
    else:
        if platform.system() == 'Windows':
            project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
            cos_path = os.path.join(project_dir, PROJECT_NAME)

            if not os.path.exists(cos_path):
                os.makedirs(cos_path)
                
        else:
            cos_path = os.path.join('/home', 'ec2-user', 'Cos_Pricing', 'Output')

        if not os.path.exists(cos_path):
            os.makedirs(cos_path)
        parts_file = os.path.join(cos_path, "Retail_Pricing_Parts_List" + "(" + str(JOB_NUMBER) + ").xlsx")

    print("Creating Retail Pricing Summary...")
    generate_retail_pricing_summary(parts_file)
    
    # Start by opening the spreadsheet and selecting the main sheet
    # try:
    #     import openpyxl
    #     import et_xmlfile
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)

    wb = openpyxl.load_workbook(parts_file)
    sheet1 = None
    sheet2 = None
    sheet3 = None
    sheet4 = None
    sheet5 = None
    sheet6 = None

    if len(MATERIAL_PARTS_LIST) != 0:
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(
            ["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
             "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE", "EDGEBANDING", "DRILLING", "PULL_DRILLING", "ASSEMBLED"])

        for i in range(len(MATERIAL_PARTS_LIST)):
            room_name = MATERIAL_PARTS_LIST[i][0]
            wall_name = MATERIAL_PARTS_LIST[i][17]
            if wall_name == "NA":
                wall_name = "---"
            sku_num = MATERIAL_PARTS_LIST[i][1]
            vendor_name = MATERIAL_PARTS_LIST[i][2]
            vendor_item = MATERIAL_PARTS_LIST[i][3]
            part_id = MATERIAL_PARTS_LIST[i][4]
            material = MATERIAL_PARTS_LIST[i][5]
            part_name = MATERIAL_PARTS_LIST[i][6]
            quantity = MATERIAL_PARTS_LIST[i][7]
            length = MATERIAL_PARTS_LIST[i][8]
            width = MATERIAL_PARTS_LIST[i][9]
            thickness = MATERIAL_PARTS_LIST[i][10]
            retail_price = MATERIAL_PARTS_LIST[i][11]
            r_labor = MATERIAL_PARTS_LIST[i][12]
            uom = MATERIAL_PARTS_LIST[i][13]
            edgebanding = MATERIAL_PARTS_LIST[i][14]
            drilling = MATERIAL_PARTS_LIST[i][18]
            pull_drilling = MATERIAL_PARTS_LIST[i][19]
            box_assembled = MATERIAL_PARTS_LIST[i][20]

            sheet1["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet1["B" + str((i + 1) + 1)] = wall_name                                         #WALL_NAME
            sheet1["C" + str((i + 1) + 1)] = sku_num                                          #SKU_NUMBER 
            sheet1["D" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet1["E" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet1["F" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            # Check to see if the next LabelID matches the current.
            # Add EB name within Melamine name if the EB name is different
            if part_id == MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][4]:
                if 'PM' in sku_num[:2] or 'BB' in sku_num[:2] and material not in MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5]:
                    sheet1["G" + str((i + 1) + 1)] = material + " (" + MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5] + ")"
                elif 'VN' in sku_num[:2] or 'WD' in sku_num[:2] and material not in MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5]:
                    sheet1["G" + str((i + 1) + 1)] = material + " (" + MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5] + ")"
                else:
                    sheet1["G" + str((i + 1) + 1)] = material                                  #MATERIAL
            else:
                sheet1["G" + str((i + 1) + 1)] = material
            sheet1["H" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet1["I" + str((i + 1) + 1)] = quantity                                          #QUANTITY
            sheet1["J" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS           
            sheet1["K" + str((i + 1) + 1)] = thickness                                         #THICKNESS
            sheet1["N" + str((i + 1) + 1)] = float(retail_price)                                  #PART_PRICE
            sheet1["O" + str((i + 1) + 1)] = float(r_labor)                                  #LABOR

            if sku_num[:2] in material_types:
                if 'SF' in uom:
                    if part_name == "Flat Crown":
                        sheet1["L" + str((i + 1) + 1)] = get_square_footage(float(length), float(width))
                        sheet1["M" + str((i + 1) + 1)] = get_linear_footage(float(length))
                    else:
                        sheet1["L" + str((i + 1) + 1)] = get_square_footage(float(length), float(width))
                        sheet1["M" + str((i + 1) + 1)] = 0

                    is_vn_sku = 'VN' in sku_num[:2]
                    is_wd_sku = 'WD' in sku_num[:2]
                    is_pbi_sku = 'pbi' in material.lower()

                    # TOTAL_PRICE
                    if is_vn_sku or is_wd_sku and not is_pbi_sku:
                        sheet1["P" + str((i + 1) + 1)] = (float(retail_price) * int(quantity))
                    elif part_name == "Flat Crown":
                        retail_price = (float(length) / 12) * float(17.00)
                        price = retail_price
                        sheet1["P" + str((i + 1) + 1)] = price
                    else:
                        price = (float(retail_price) * int(quantity)) * get_square_footage(float(length), float(width))
                        labor = float(r_labor) * int(quantity)
                        sheet1["P" + str((i + 1) + 1)] = price + labor

                    if len(EDGEBANDING) > 0:
                        for index, sublist in enumerate(EDGEBANDING):
                            if sublist[0] == part_id:
                                if EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = "--"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 1:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                elif EDGEBANDING[index][1] == 1 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                elif EDGEBANDING[index][1] == 2 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 2:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                else:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S_" + str(EDGEBANDING[index][2]) + "L"
                if 'LF' in uom:
                    is_bb_sku = 'BB' in sku_num[:2]

                    if is_bb_sku:
                        sheet1["L" + str((i + 1) + 1)] = 0
                        sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(float(length)))
                        sheet1["P" + str((i + 1) + 1)] = (float(retail_price) * int(quantity)) * get_linear_footage(float(length))   #TOTAL_PRICE

                    if edgebanding is not None:
                        eb_length = get_eb_measurements(edgebanding, float(length), float(width))
                        if eb_length:
                            sheet1["L" + str((i + 1) + 1)] = 0
                            sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(eb_length)) + " (" + str(edgebanding) + ")"
                            sheet1["P" + str((i + 1) + 1)] = (float(retail_price) * int(quantity)) * get_linear_footage(eb_length)   #TOTAL_PRICE
                    if len(EDGEBANDING) > 0:
                        for index, sublist in enumerate(EDGEBANDING):
                            if sublist[0] == part_id:
                                if EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = "--"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 1:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                elif EDGEBANDING[index][1] == 1 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                else:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S_" + str(EDGEBANDING[index][2]) + "L"
            sheet1["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["P" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

            sheet1["R" + str((i + 1) + 1)] = drilling
            sheet1["S" + str((i + 1) + 1)] = pull_drilling
            sheet1["T" + str((i + 1) + 1)] = box_assembled

        row_max = sheet1.max_row
        sheet1["R" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet1["R" + str(row_max + 3)] = "TOTAL"
        sheet1["R" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet1["R" + str(row_max + 4)] = "=SUM(P2:P" + str(row_max) + ")"

        sheet1.column_dimensions['Q'].hidden = True
        sheet1.column_dimensions['O'].hidden = True
        set_column_width(sheet1)
    else:
        print("Materials Parts List Empty")
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE", "EDGEBANDING"])

        # sheet1["A2"] = "---"                                          #ROOM_NAME
        # sheet1["B2"] = "---"                                          #WALL_NAME
        # sheet1["C2"] = "---"                                          #SKU_NUMBER 
        # sheet1["D2"] = "---"                                          #VENDOR_NAME
        # sheet1["E2"] = "---"                                          #VENDOR_ITEM
        # sheet1["F2"] = "---"                                          #PART LABELID
        # sheet1["G2"] = "---"                                          #MATERIAL
        # sheet1["H2"] = "---"                                          #PART_NAME    
        # sheet1["I2"] = 0                                              #QUANTITY
        # sheet1["J2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet1["K2"] = "---"                                          #THICKNESS
        # sheet1["L2"] = 0                                           
        # sheet1["M2"] = 0
        # sheet1["N2"] = 0                                              #PART_PRICE
        # sheet1["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["O2"] = 0                                              #LABOR
        # sheet1["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["P2"] = 0                                              #TOTAL_PRICE
        # sheet1["P2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["Q2"] = "---"

        sheet1.column_dimensions['Q'].hidden = True
        sheet1.column_dimensions['O'].hidden = True
        set_column_width(sheet1)

    if len(HARDWARE_PARTS_LIST) != 0:
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])
        
        for i in range(len(HARDWARE_PARTS_LIST)):
            room_name = HARDWARE_PARTS_LIST[i][0]         
            sku_num = HARDWARE_PARTS_LIST[i][1]
            vendor_name = HARDWARE_PARTS_LIST[i][2]           
            vendor_item = HARDWARE_PARTS_LIST[i][3]
            part_id = HARDWARE_PARTS_LIST[i][4] 
            part_name = HARDWARE_PARTS_LIST[i][5]
            quantity = HARDWARE_PARTS_LIST[i][6]
            retail_price = HARDWARE_PARTS_LIST[i][7]

            sheet2["A" + str((i + 1) + 1)] = room_name         
            sheet2["B" + str((i + 1) + 1)] = sku_num
            sheet2["C" + str((i + 1) + 1)] = vendor_name           
            sheet2["D" + str((i + 1) + 1)] = vendor_item
            sheet2["E" + str((i + 1) + 1)] = part_id 
            sheet2["F" + str((i + 1) + 1)] = part_name
            sheet2["H" + str((i + 1) + 1)] = quantity
            sheet2["I" + str((i + 1) + 1)] = float(retail_price)
            sheet2["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet2["J" + str((i + 1) + 1)] = float(retail_price) * int(quantity)
            sheet2["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet2.max_row
        sheet2["K" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet2["K" + str(row_max + 3)] = "TOTAL"
        sheet2["K" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet2["K" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet2)
    else:
        print("Hardware Parts List Empty")
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        # sheet2["A2"] = "---"                                          #ROOM_NAME
        # sheet2["B2"] = "---"                                          #SKU_NUMBER
        # sheet2["C2"] = "---"                                          #VENDOR_NAME 
        # sheet2["D2"] = "---"                                          #VENDOR_ITEM
        # sheet2["E2"] = "---"                                          #PART LABELID
        # sheet2["F2"] = "---"                                          #PART_NAME
        # sheet2["H2"] = 0                                          #QUANTITY   
        # sheet2["I2"] = 0                                              #PART_PRICE
        # sheet2["I2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet2["J2"] = 0                                              #TOTAL_PRICE
        # sheet2["J2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet2["Q2"] = "---"

        set_column_width(sheet2)

    if len(ACCESSORY_PARTS_LIST) != 0:
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "PART_DIMENSIONS", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        for i in range(len(ACCESSORY_PARTS_LIST)):
            room_name = ACCESSORY_PARTS_LIST[i][0]         
            sku_num = ACCESSORY_PARTS_LIST[i][1]
            vendor_name = ACCESSORY_PARTS_LIST[i][2]           
            vendor_item = ACCESSORY_PARTS_LIST[i][3]
            part_id = ACCESSORY_PARTS_LIST[i][4] 
            part_name = ACCESSORY_PARTS_LIST[i][5]
            length = ACCESSORY_PARTS_LIST[i][6]
            width = ACCESSORY_PARTS_LIST[i][7]
            quantity = ACCESSORY_PARTS_LIST[i][8]
            retail_price = ACCESSORY_PARTS_LIST[i][9]
            uom = ACCESSORY_PARTS_LIST[i][11]

            sheet3["A" + str((i + 1) + 1)] = room_name     
            sheet3["B" + str((i + 1) + 1)] = sku_num
            sheet3["C" + str((i + 1) + 1)] = vendor_name           
            sheet3["D" + str((i + 1) + 1)] = vendor_item
            sheet3["E" + str((i + 1) + 1)] = part_id           
            sheet3["F" + str((i + 1) + 1)] = part_name
            if "SF" in uom:
                sheet3["G" + str((i + 1) + 1)] = str(length) + " x " + str(width) + " (" + str(get_square_footage(float(length), float(width))) + " SF)"
            elif "LF" in uom:
                sheet3["G" + str((i + 1) + 1)] = str(length) + " (" + str(get_linear_footage(float(length))) + " LF)"
            else:
                sheet3["G" + str((i + 1) + 1)] = 0
            sheet3["H" + str((i + 1) + 1)] = quantity
            sheet3["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            if "SF" in uom:
                sheet3["I" + str((i + 1) + 1)] = float(retail_price) * get_square_footage(float(length), float(width))
            elif "LF" in uom:
                sheet3["I" + str((i + 1) + 1)] = float(retail_price) * get_linear_footage(float(length))
            else:
                sheet3["I" + str((i + 1) + 1)] = float(retail_price)
            
            sheet3["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            if "SF" in uom:
                sheet3["J" + str((i + 1) + 1)] = (float(retail_price) * int(quantity)) * get_square_footage(float(length), float(width))
            elif "LF" in uom:
                sheet3["J" + str((i + 1) + 1)] = (float(retail_price) * int(quantity)) * get_linear_footage(float(length))
            else:
                sheet3["J" + str((i + 1) + 1)] = float(retail_price) * int(quantity)

        row_max = sheet3.max_row
        sheet3["K" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet3["K" + str(row_max + 3)] = "TOTAL"
        sheet3["K" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet3["K" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet3)
    else:
        print("Accessory Parts List Empty")
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "PART_DIMENSIONS", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        # sheet3["A2"] = "---"                                          #ROOM_NAME
        # sheet3["B2"] = "---"                                          #SKU_NUMBER
        # sheet3["C2"] = "---"                                          #VENDOR_NAME 
        # sheet3["D2"] = "---"                                          #VENDOR_ITEM
        # sheet3["E2"] = "---"                                          #PART LABELID
        # sheet3["F2"] = "---"                                          #PART_NAME
        # sheet3["G2"] = "---"                                          #PART_DIMENSIONS
        # sheet3["H2"] = 0                                              #QUANTITY   
        # sheet3["I2"] = 0                                              #LABOR
        # sheet3["I2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet3["J2"] = 0                                              #TOTAL_PRICE
        # sheet3["J2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet3["Q2"] = "---"

        set_column_width(sheet3)

    if len(UPGRADED_PANEL_PARTS_LIST) != 0:
        sheet4 = wb.create_sheet()
        sheet4.title = "Upgraded Panels"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PAINT_STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        for i in range(len(UPGRADED_PANEL_PARTS_LIST)):
            room_name = UPGRADED_PANEL_PARTS_LIST[i][0]
            sku_num = UPGRADED_PANEL_PARTS_LIST[i][1]
            vendor_name = UPGRADED_PANEL_PARTS_LIST[i][2]
            vendor_item = UPGRADED_PANEL_PARTS_LIST[i][3]
            part_id = UPGRADED_PANEL_PARTS_LIST[i][4]
            paint_stain = UPGRADED_PANEL_PARTS_LIST[i][5]
            part_name = UPGRADED_PANEL_PARTS_LIST[i][6]
            quantity = UPGRADED_PANEL_PARTS_LIST[i][7]
            length = UPGRADED_PANEL_PARTS_LIST[i][8]
            width = UPGRADED_PANEL_PARTS_LIST[i][9]    
            thickness = UPGRADED_PANEL_PARTS_LIST[i][10]
            retail_price = UPGRADED_PANEL_PARTS_LIST[i][11]
            r_labor = UPGRADED_PANEL_PARTS_LIST[i][12]

            sheet4["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet4["B" + str((i + 1) + 1)] = sku_num                                          #SKU_NUMBER 
            sheet4["C" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet4["D" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet4["E" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            sheet4["F" + str((i + 1) + 1)] = paint_stain                                          #PAINT_STAIN_COLOR
            sheet4["G" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet4["H" + str((i + 1) + 1)] = quantity                                          #QUANTITY
            sheet4["I" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS
            sheet4["J" + str((i + 1) + 1)] = thickness                                   #THICKNESS

            # if 'Glass' in part_name and 'Glass Shelf' not in part_name:
            #     if UPGRADED_PANEL_PARTS_LIST[i][15] is not None:
            #         if UPGRADED_PANEL_PARTS_LIST[i][15] == 'Yes':
            #             sheet4["K" + str((i + 1) + 1)] = get_square_footage(round((float(length) - 6.75)/2, 2),round(float(width) - 4.75, 2)) * int(quantity)
            #         else:
            #             sheet4["K" + str((i + 1) + 1)] = get_square_footage(round(float(length) - 4.75, 2),round(float(width) - 4.75, 2)) * int(quantity)        
            # else:
            sheet4["K" + str((i + 1) + 1)] = get_square_footage(float(length),float(width)) * int(quantity)
            
            sheet4["L" + str((i + 1) + 1)] = "---"
            sheet4["M" + str((i + 1) + 1)] = float(retail_price)                                  #PART_PRICE
            sheet4["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["N" + str((i + 1) + 1)] = float(r_labor)                                  #LABOR
            sheet4["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["O" + str((i + 1) + 1)] = float(retail_price)  #TOTAL_PRICE
            sheet4["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet4.max_row
        sheet4["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet4["P" + str(row_max + 3)] = "TOTAL"
        sheet4["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet4["P" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet4.column_dimensions['N'].hidden = True
        set_column_width(sheet4)
    else:
        print("Upgraded Panels Parts List Empty")
        sheet4 = wb.create_sheet()
        sheet4.title = "Upgraded Panels"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PAINT_STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        # sheet4["A2"] = "---"                                          #ROOM_NAME
        # sheet4["B2"] = "---"                                          #SKU_NUMBER 
        # sheet4["C2"] = "---"                                          #VENDOR_NAME
        # sheet4["D2"] = "---"                                          #VENDOR_ITEM
        # sheet4["E2"] = "---"                                          #PART LABELID
        # sheet4["F2"] = "---"                                          #MATERIAL
        # sheet4["G2"] = "---"                                          #PART_NAME    
        # sheet4["H2"] = 0                                              #QUANTITY
        # sheet4["I2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet4["J2"] = "---"                                          #THICKNESS
        # sheet4["K2"] = "---"
        # sheet4["L2"] = "---"
        # sheet4["M2"] = 0                                              #PART_PRICE
        # sheet4["M2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet4["N2"] = 0                                              #LABOR
        # sheet4["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet4["O2"] = 0                                              #TOTAL_PRICE
        # sheet4["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        sheet4.column_dimensions['N'].hidden = True
        set_column_width(sheet4)

    if len(GLASS_PARTS_LIST) != 0:
        sheet5 = wb.create_sheet()
        sheet5.title = "Glass"
        sheet5.append([
            "ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID",
            "MATERIAL", "PART_NAME", "QUANTITY", "PART_DIMENSIONS", "THICKNESS",
            "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        for i in range(len(GLASS_PARTS_LIST)):
            room_name = GLASS_PARTS_LIST[i][0]
            sku_num = GLASS_PARTS_LIST[i][1]
            vendor_name = GLASS_PARTS_LIST[i][2]
            vendor_item = GLASS_PARTS_LIST[i][3]
            part_id = GLASS_PARTS_LIST[i][4]
            material = GLASS_PARTS_LIST[i][5]
            part_name = GLASS_PARTS_LIST[i][6]
            quantity = GLASS_PARTS_LIST[i][7]
            length = GLASS_PARTS_LIST[i][8]
            width = GLASS_PARTS_LIST[i][9]
            thickness = GLASS_PARTS_LIST[i][10]
            retail_price = GLASS_PARTS_LIST[i][11]
            r_labor = GLASS_PARTS_LIST[i][12]

            is_glass = 'Glass' in part_name and 'Glass Shelf' not in part_name
            if is_glass and GLASS_PARTS_LIST[i][15] == 'Yes':
                quantity = int(quantity) * 2

            if sku_num == "GL-0000010":
                thickness = 0.125
                quantity = int(quantity) * 2

            sheet5["A" + str((i + 1) + 1)] = room_name        # ROOM_NAME
            sheet5["B" + str((i + 1) + 1)] = sku_num          # SKU_NUMBER
            sheet5["C" + str((i + 1) + 1)] = vendor_name      # VENDOR_NAME
            sheet5["D" + str((i + 1) + 1)] = vendor_item      # VENDOR_ITEM
            sheet5["E" + str((i + 1) + 1)] = part_id          # PART LABELID
            sheet5["F" + str((i + 1) + 1)] = material         # MATERIAL
            sheet5["G" + str((i + 1) + 1)] = part_name        # PART_NAME
            sheet5["H" + str((i + 1) + 1)] = int(quantity)    # QUANTITY

            # Normalize indices for readability
            row_idx = i + 2  # (i + 1) + 1
            col_i = "I" + str(row_idx)
            col_j = "J" + str(row_idx)
            col_k = "K" + str(row_idx)

            # Check for glass part
            is_glass_shelf = 'Glass Shelf' in part_name
            is_glass = 'Glass' in part_name
            square_footage = None

            if is_glass_shelf:
                sheet5[col_i] = f"{length} x {width}"  # PART_DIMENSIONS
            elif is_glass:
                style_name = GLASS_PARTS_LIST[i][16]

                # If door has center rail
                if is_glass and GLASS_PARTS_LIST[i][15] == 'Yes':
                    adjusted_length, adjusted_width = get_glass_inset_size(length, width, True, style_name)
                    square_footage = get_square_footage(adjusted_length, adjusted_width)
                else:
                    adjusted_length, adjusted_width = get_glass_inset_size(length, width, False, style_name)
                    square_footage = get_square_footage(adjusted_length, adjusted_width)

                # Write to sheet
                sheet5[col_i] = f"{adjusted_length} x {adjusted_width}"  # PART_DIMENSIONS

            sheet5[col_j] = thickness  # THICKNESS
            if square_footage is not None:
                sheet5[col_k] = square_footage
            else:
                sheet5["K" + str((i + 1) + 1)] = get_square_footage(float(length), float(width))

            sheet5["L" + str((i + 1) + 1)] = get_linear_footage(float(length))
            sheet5["M" + str((i + 1) + 1)] = float(retail_price)                                  #PART_PRICE
            sheet5["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["N" + str((i + 1) + 1)] = float(r_labor)                                  #LABOR
            sheet5["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["O" + str((i + 1) + 1)] = float(retail_price) * int(quantity)
            sheet5["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet5.max_row
        sheet5["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet5["P" + str(row_max + 3)] = "TOTAL"
        sheet5["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet5["P" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet5.column_dimensions['N'].hidden = True
        set_column_width(sheet5)
    else:
        print("Glass Parts List Empty")
        sheet5 = wb.create_sheet()
        sheet5.title = "Glass"
        sheet5.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])


        # sheet5["A2"] = "---"                                          #ROOM_NAME
        # sheet5["B2"] = "---"                                          #SKU_NUMBER 
        # sheet5["C2"] = "---"                                          #VENDOR_NAME
        # sheet5["D2"] = "---"                                          #VENDOR_ITEM
        # sheet5["E2"] = "---"                                          #PART LABELID
        # sheet5["F2"] = "---"                                          #MATERIAL
        # sheet5["G2"] = "---"                                          #PART_NAME    
        # sheet5["H2"] = 0                                              #QUANTITY
        # sheet5["I2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet5["J2"] = "---"                                          #THICKNESS
        # sheet5["K2"] = 0
        # sheet5["L2"] = 0
        # sheet5["M2"] = 0                                              #PART_PRICE
        # sheet5["M2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet5["N2"] = 0                                              #LABOR
        # sheet5["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet5["O2"] = 0                                              #TOTAL_PRICE
        # sheet5["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        sheet5.column_dimensions['N'].hidden = True
        set_column_width(sheet5)


    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        sheet6 = wb.create_sheet()
        sheet6.title = "Special Order"
        sheet6.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])
        
        sheet6["Q1"] = "Be sure to note ALL material changes and special order additions on your plans"
        sheet6["Q1"].fill = openpyxl.styles.PatternFill("solid", start_color="FFFF00")
        sheet6["Q1"].font = openpyxl.styles.Font(bold=True)
        sheet6["Q1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # sheet6.merge_cells(start_row=2, start_column=17, end_row=9, end_column=17)
        # cell = sheet6.cell(row=2, column=17)
        # cell.value = "Be sure to note ALL material changes and special order additions on your plans"
        # cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # cell.fill = openpyxl.styles.PatternFill("solid", start_color="FFFF00")
        # cell.font = openpyxl.styles.Font(bold=True)

        for i in range(len(SPECIAL_ORDER_PARTS_LIST)):
            room_name = SPECIAL_ORDER_PARTS_LIST[i][0]
            sku_num = SPECIAL_ORDER_PARTS_LIST[i][1]
            vendor_name = SPECIAL_ORDER_PARTS_LIST[i][2]
            vendor_item = SPECIAL_ORDER_PARTS_LIST[i][3]
            part_id = SPECIAL_ORDER_PARTS_LIST[i][4]
            material = SPECIAL_ORDER_PARTS_LIST[i][5]
            part_name = SPECIAL_ORDER_PARTS_LIST[i][6]
            quantity = SPECIAL_ORDER_PARTS_LIST[i][7]
            length = SPECIAL_ORDER_PARTS_LIST[i][8]
            width = SPECIAL_ORDER_PARTS_LIST[i][9]    
            thickness = SPECIAL_ORDER_PARTS_LIST[i][10]
            retail_price = SPECIAL_ORDER_PARTS_LIST[i][11]
            r_labor = SPECIAL_ORDER_PARTS_LIST[i][12]

            sheet6["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet6["B" + str((i + 1) + 1)] = sku_num                                          #SKU_NUMBER 
            sheet6["C" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet6["D" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet6["E" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            sheet6["F" + str((i + 1) + 1)] = material                                          #MATERIAL
            sheet6["G" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet6["H" + str((i + 1) + 1)] = int(quantity)                                         #QUANTITY
            if 'Glass' in part_name and 'Glass Shelf' not in part_name:
                if SPECIAL_ORDER_PARTS_LIST[i][15] is not None:
                    if SPECIAL_ORDER_PARTS_LIST[i][15] == 'Yes':
                        sheet6["I" + str((i + 1) + 1)] = str(round((float(length) - 6.75)/2, 2)) + " x " + str(round(float(width) - 4.75, 2))
                    else:
                        sheet6["I" + str((i + 1) + 1)] = str(round(float(length) - 4.75, 2)) + " x " + str(round(float(width) - 4.75, 2))      #PART_DIMENSIONS
            else:
                sheet6["I" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS           
            sheet6["J" + str((i + 1) + 1)] = thickness                                         #THICKNESS
            if 'Glass' in part_name and 'Glass Shelf' not in part_name:
                if SPECIAL_ORDER_PARTS_LIST[i][15] is not None:
                    if SPECIAL_ORDER_PARTS_LIST[i][15] == 'Yes':
                        sheet6["K" + str((i + 1) + 1)] = get_square_footage(round((float(length) - 6.75)/2, 2),round(float(width) - 4.75, 2)) * int(quantity)
                    else:
                        sheet6["K" + str((i + 1) + 1)] = get_square_footage(round(float(length) - 4.75, 2),round(float(width) - 4.75, 2)) * int(quantity)
                        
            else:
                sheet6["K" + str((i + 1) + 1)] = get_square_footage(float(length),float(width)) * int(quantity)
            sheet6["L" + str((i + 1) + 1)] = get_linear_footage(float(length))
            sheet6["M" + str((i + 1) + 1)] = float(retail_price)                                  #PART_PRICE
            sheet6["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet6["N" + str((i + 1) + 1)] = float(r_labor)                                  #LABOR
            sheet6["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet6["O" + str((i + 1) + 1)] = (float(retail_price) + float(r_labor)) * int(quantity)
            sheet6["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        # row_max = sheet6.max_row
        # sheet6["O" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        # sheet6["O" + str(row_max + 3)] = "TOTAL"
        # sheet6["O" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet6["O" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet6.column_dimensions['N'].hidden = True
        set_column_width(sheet6)
    else:
        print("Special Order Parts List Empty")
        sheet6 = wb.create_sheet()
        sheet6.title = "Special Order"
        sheet6.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])
        
        sheet6["Q1"] = "Be sure to note ALL material changes and special order additions on your plans"
        sheet6["Q1"].fill = openpyxl.styles.PatternFill("solid", start_color="FFFF00")
        sheet6["Q1"].font = openpyxl.styles.Font(bold=True)
        sheet6["Q1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet6.column_dimensions['N'].hidden = True
        set_column_width(sheet6)

    # Save the Workbook
    try:
        wb.save(filename=parts_file)
        print("Retail Pricing Parts List Generated")
        print("Creating Retail Parts Summary...")
    except PermissionError:
        parts_file_name = os.path.basename(parts_file)
        message = "Cannot create parts list. Pricing Spreadsheet Open.\nPlease close the file:\n" + parts_file_name
        return bpy.ops.snap.message_box('INVOKE_DEFAULT', message=message, icon='ERROR')

    # Create Pivot table for Materials and Glass Square Footage on Summary Sheet
    df_materials = pandas.read_excel(
        parts_file,
        sheet_name='Materials'
    ).query(
        'SKU_NUMBER.str.contains("EB")==False and MATERIAL.str.contains("BBBB")==False',
        engine='python'
    )

    # sf_materials = pandas.pivot_table(df_materials, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT'], aggfunc=numpy.sum)

    # Calculate and add column for total square footage
    df_materials['SQUARE_FT_TOTAL'] = df_materials['QUANTITY'] * df_materials['SQUARE_FT']

    # Aggregate square footage total
    sf_materials = pandas.pivot_table(df_materials, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT_TOTAL'], aggfunc=numpy.sum)

    # Rename SQUARE_FT_TOTAL column
    sf_materials = sf_materials.rename(columns={'SQUARE_FT_TOTAL': 'SQUARE_FT'})    

    df_glass = pandas.read_excel(parts_file, sheet_name='Glass')
    sf_glass = pandas.pivot_table(df_glass, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT'], aggfunc=numpy.sum)

    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                                                right=openpyxl.styles.borders.Side(style='thin'), 
                                                top=openpyxl.styles.borders.Side(style='thin'), 
                                                bottom=openpyxl.styles.borders.Side(style='thin'))

    with pandas.ExcelWriter(parts_file, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        sf_row_start = 0

        if not sf_materials.empty and not sf_glass.empty:
            frames = [df_materials, df_glass]
            all_frames = pandas.concat(frames)
            all_data = pandas.pivot_table(all_frames, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT'], aggfunc=numpy.sum)
            all_data.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=2)
            sf_row_start = (all_data.shape[0] + 1) + 4
            room_data = pandas.pivot_table(all_frames, index=['ROOM_NAME'], values=['SQUARE_FT'], aggfunc=numpy.sum)
            room_data.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=sf_row_start)
            sf_row_start = (all_data.shape[0] + 1) + (room_data.shape[0] + 1) + 4
            project_data = all_frames['SQUARE_FT'].sum()
            ws = writer.sheets["Retail Pricing Summary"]
            ws["D" + str(sf_row_start+3)] = "Project Materials SF Total"
            ws["D" + str(sf_row_start+3)].font = openpyxl.styles.Font(bold=True)
            ws["D" + str(sf_row_start+3)].border = thin_border
            ws["E" + str(sf_row_start+3)] = project_data

        else:
            if not sf_materials.empty:
                # Calculate total square footage per room
                room_materials = df_materials.groupby('ROOM_NAME').apply(calc_room_total_sq_ft).reset_index(name='SQUARE_FT')

                sf_materials.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=2)
                sf_row_start = (sf_materials.shape[0] + 1) + 4

                # Write without the default index column
                room_materials.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=sf_row_start, index=False)

                # Apply borders to room names column
                for i in range(len(room_materials) + 1):
                    ws = writer.sheets["Retail Pricing Summary"]
                    ws["D" + str(sf_row_start + i + 1)].border = thin_border

                sf_row_start = (sf_materials.shape[0] + 1) + (room_materials.shape[0] + 1) + 4

                # Calculate total square footage for project
                total_sq_ft = (df_materials['QUANTITY'] * df_materials['SQUARE_FT']).sum()
                ws = writer.sheets["Retail Pricing Summary"]
                ws["D" + str(sf_row_start + 3)] = "Project Material SF Total"
                ws["D" + str(sf_row_start + 3)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start + 3)].border = thin_border
                ws["E" + str(sf_row_start + 3)] = total_sq_ft

            if not sf_glass.empty:
                sf_glass.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=2)
                sf_row_start = (sf_glass.shape[0] + 1) + 4
                room_glass = pandas.pivot_table(df_glass, index=['ROOM_NAME'], values=['SQUARE_FT'], aggfunc=numpy.sum)
                room_glass.to_excel(writer, 'Retail Pricing Summary', startcol=3, startrow=sf_row_start)
                sf_row_start = (sf_glass.shape[0] + 1) + (room_glass.shape[0] + 1) + 4
                glass_total_sq_ft = (df_glass['QUANTITY'] * df_glass['SQUARE_FT']).sum()
                ws = writer.sheets["Retail Pricing Summary"]
                ws["D" + str(sf_row_start+3)] = "Project Glass SF Total"
                ws["D" + str(sf_row_start+3)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start+3)].border = thin_border
                ws["E" + str(sf_row_start+3)] = glass_total_sq_ft
                
        set_column_width(writer.sheets['Retail Pricing Summary'])

    wb.save(filename=parts_file)
    generate_parts_summary(parts_file, sheet1, sheet2, sheet3, sheet4, sheet5, sheet6)


def generate_franchise_parts_list():
    if not COS_FLAG:
        props = bpy.context.window_manager.sn_project
        proj = props.get_project()
        cleaned_name = proj.get_clean_name(proj.name)
        project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
        selected_project = os.path.join(project_dir, cleaned_name)
        parts_file = os.path.join(selected_project, "Franchise_Pricing_Parts_List" + "(" + str(cleaned_name) + ").xlsx")
    
        if not os.path.exists(project_dir):
            print("Projects Directory does not exist")
    else:
        if platform.system() == 'Windows':
            project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
            cos_path = os.path.join(project_dir, PROJECT_NAME)

            if not os.path.exists(cos_path):
                os.makedirs(cos_path)
        else:
            cos_path = os.path.join('/home', 'ec2-user', 'Cos_Pricing', 'Output')

        if not os.path.exists(cos_path):
            os.makedirs(cos_path)
        parts_file = os.path.join(cos_path, "Franchise_Pricing_Parts_List" + "(" + str(JOB_NUMBER) + ").xlsx")
        

    print("Creating Franchise Pricing Summary...")
    generate_franchise_pricing_summary(parts_file)
    
    # Start by opening the spreadsheet and selecting the main sheet
    # try:
    #     import openpyxl
    #     import et_xmlfile
    # except ModuleNotFoundError:
    #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
    #     sys.path.append(python_lib_path)

    wb = openpyxl.load_workbook(parts_file)
    sheet1 = None
    sheet2 = None
    sheet3 = None
    sheet4 = None
    sheet5 = None
    sheet6 = None

    if len(MATERIAL_PARTS_LIST) != 0:
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE", "EDGEBANDING", "DRILLING", "PULL_DRILLING", "ASSEMBLED"])

        for i in range(len(MATERIAL_PARTS_LIST)):
            room_name = MATERIAL_PARTS_LIST[i][0]
            wall_name = MATERIAL_PARTS_LIST[i][17]
            if wall_name == "NA":
                wall_name = "---"
            sku_num = MATERIAL_PARTS_LIST[i][1]
            vendor_name = MATERIAL_PARTS_LIST[i][2]
            vendor_item = MATERIAL_PARTS_LIST[i][3]
            part_id = MATERIAL_PARTS_LIST[i][4]
            material = MATERIAL_PARTS_LIST[i][5]
            part_name = MATERIAL_PARTS_LIST[i][6]
            quantity = MATERIAL_PARTS_LIST[i][7]
            length = MATERIAL_PARTS_LIST[i][8]
            width = MATERIAL_PARTS_LIST[i][9]
            thickness = MATERIAL_PARTS_LIST[i][10]
            franchise_price = MATERIAL_PARTS_LIST[i][16]
            f_labor = MATERIAL_PARTS_LIST[i][15]
            uom = MATERIAL_PARTS_LIST[i][13]
            edgebanding = MATERIAL_PARTS_LIST[i][14]
            drilling = MATERIAL_PARTS_LIST[i][18]
            pull_drilling = MATERIAL_PARTS_LIST[i][19]
            box_assembled = MATERIAL_PARTS_LIST[i][20]

            sheet1["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet1["B" + str((i + 1) + 1)] = wall_name                                         #WALL_NAME
            sheet1["C" + str((i + 1) + 1)] = sku_num                                          #SKU_NUMBER 
            sheet1["D" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet1["E" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet1["F" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            if part_id == MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][4]:
                if 'PM' in sku_num[:2] or 'BB' in sku_num[:2] and material not in MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5]:
                    sheet1["G" + str((i + 1) + 1)] = material + " (" + MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5] + ")"
                elif 'VN' in sku_num[:2] or 'WD' in sku_num[:2] and material not in MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5]:
                    sheet1["G" + str((i + 1) + 1)] = material + " (" + MATERIAL_PARTS_LIST[(i+1)%len(MATERIAL_PARTS_LIST)][5] + ")"
                else:
                    sheet1["G" + str((i + 1) + 1)] = material                                  #MATERIAL
            else:
                sheet1["G" + str((i + 1) + 1)] = material
            sheet1["H" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet1["I" + str((i + 1) + 1)] = quantity                                          #QUANTITY
            sheet1["J" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS           
            sheet1["K" + str((i + 1) + 1)] = thickness                                         #THICKNESS
            sheet1["N" + str((i + 1) + 1)] = float(franchise_price)                                  #PART_PRICE
            sheet1["O" + str((i + 1) + 1)] = float(f_labor)                                  #LABOR

            if sku_num[:2] in material_types:
                if 'SF' in uom:
                    sheet1["L" + str((i + 1) + 1)] = get_square_footage(float(length), float(width))
                    sheet1["M" + str((i + 1) + 1)] = 0

                    is_vn_sku = 'VN' in sku_num[:2]
                    is_wd_sku = 'WD' in sku_num[:2]
                    is_pbi_sku = 'pbi' in material.lower()

                    # TOTAL_PRICE
                    if is_vn_sku or is_wd_sku and not is_pbi_sku:
                        sheet1["P" + str((i + 1) + 1)] = (float(franchise_price) * int(quantity))
                    else:
                        sqft = get_square_footage(float(length), float(width))
                        sqft_price = float(franchise_price) * sqft
                        price = sqft_price * float(quantity)
                        labor = float(f_labor) * int(quantity)
                        sheet1["P" + str((i + 1) + 1)] = price + labor

                    if len(EDGEBANDING) > 0:
                        for index, sublist in enumerate(EDGEBANDING):
                            if sublist[0] == part_id:
                                if EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = "--"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 1:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                elif EDGEBANDING[index][1] == 1 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                elif EDGEBANDING[index][1] == 2 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 2:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                else:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S_" + str(EDGEBANDING[index][2]) + "L"
    
                if 'LF' in uom:
                    is_bb_sku = 'BB' in sku_num[:2]

                    if is_bb_sku:
                        sheet1["L" + str((i + 1) + 1)] = 0
                        sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(float(length)))
                        sheet1["P" + str((i + 1) + 1)] = (float(franchise_price) * int(quantity)) * get_linear_footage(float(length))   #TOTAL_PRICE

                    if edgebanding is not None:
                        eb_length = get_eb_measurements(edgebanding, float(length), float(width))
                        if eb_length:
                            sheet1["L" + str((i + 1) + 1)] = 0
                            sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(eb_length)) + " (" + str(edgebanding) + ")"
                            sheet1["P" + str((i + 1) + 1)] = (float(franchise_price) * int(quantity)) * get_linear_footage(eb_length)   #TOTAL_PRICE
                    if len(EDGEBANDING) > 0:
                        for index, sublist in enumerate(EDGEBANDING):
                            if sublist[0] == part_id:
                                if EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = "--"
                                elif EDGEBANDING[index][1] == 0 and EDGEBANDING[index][2] == 1:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][2]) + "L"
                                elif EDGEBANDING[index][1] == 1 and EDGEBANDING[index][2] == 0:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S"
                                else:
                                    sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[index][1]) + "S_" + str(EDGEBANDING[index][2]) + "L"
            sheet1["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["P" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

            sheet1["R" + str((i + 1) + 1)] = drilling
            sheet1["S" + str((i + 1) + 1)] = pull_drilling
            sheet1["T" + str((i + 1) + 1)] = box_assembled

        row_max = sheet1.max_row
        sheet1["R" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet1["R" + str(row_max + 3)] = "TOTAL"
        sheet1["R" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet1["R" + str(row_max + 4)] = "=SUM(P2:P" + str(row_max) + ")"

        sheet1.column_dimensions['Q'].hidden = True
        sheet1.column_dimensions['O'].hidden = True
        set_column_width(sheet1)
    else:
        print("Materials Parts List Empty")
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE", "EDGEBANDING"])

        # sheet1["A2"] = "---"                                          #ROOM_NAME
        # sheet1["B2"] = "---"                                          #WALL_NAME
        # sheet1["C2"] = "---"                                          #SKU_NUMBER 
        # sheet1["D2"] = "---"                                          #VENDOR_NAME
        # sheet1["E2"] = "---"                                          #VENDOR_ITEM
        # sheet1["F2"] = "---"                                          #PART LABELID
        # sheet1["G2"] = "---"                                          #MATERIAL
        # sheet1["H2"] = "---"                                          #PART_NAME    
        # sheet1["I2"] = 0                                              #QUANTITY
        # sheet1["J2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet1["K2"] = "---"                                          #THICKNESS
        # sheet1["L2"] = 0                                           
        # sheet1["M2"] = 0
        # sheet1["N2"] = 0                                              #PART_PRICE
        # sheet1["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["O2"] = 0                                              #LABOR
        # sheet1["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["P2"] = 0                                              #TOTAL_PRICE
        # sheet1["P2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet1["Q2"] = "---"

        sheet1.column_dimensions['Q'].hidden = True
        sheet1.column_dimensions['O'].hidden = True
        set_column_width(sheet1)

    if len(HARDWARE_PARTS_LIST) != 0:
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])
        
        for i in range(len(HARDWARE_PARTS_LIST)):
            room_name = HARDWARE_PARTS_LIST[i][0]         
            sku_num = HARDWARE_PARTS_LIST[i][1]
            vendor_name = HARDWARE_PARTS_LIST[i][2]           
            vendor_item = HARDWARE_PARTS_LIST[i][3]
            part_id = HARDWARE_PARTS_LIST[i][4] 
            part_name = HARDWARE_PARTS_LIST[i][5]
            quantity = HARDWARE_PARTS_LIST[i][6]
            franchise_price = HARDWARE_PARTS_LIST[i][8]

            sheet2["A" + str((i + 1) + 1)] = room_name         
            sheet2["B" + str((i + 1) + 1)] = sku_num
            sheet2["C" + str((i + 1) + 1)] = vendor_name           
            sheet2["D" + str((i + 1) + 1)] = vendor_item
            sheet2["E" + str((i + 1) + 1)] = part_id 
            sheet2["F" + str((i + 1) + 1)] = part_name
            sheet2["H" + str((i + 1) + 1)] = quantity
            sheet2["I" + str((i + 1) + 1)] = float(franchise_price)
            sheet2["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet2["J" + str((i + 1) + 1)] = float(franchise_price) * int(quantity)
            sheet2["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet2.max_row
        sheet2["K" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet2["K" + str(row_max + 3)] = "TOTAL"
        sheet2["K" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet2["K" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet2)
    else:
        print("Hardware Parts List Empty")
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        # sheet2["A2"] = "---"                                          #ROOM_NAME
        # sheet2["B2"] = "---"                                          #WALL_NAME
        # sheet2["C2"] = "---"                                          #SKU_NUMBER 
        # sheet2["D2"] = "---"                                          #VENDOR_NAME
        # sheet2["E2"] = "---"                                          #VENDOR_ITEM
        # sheet2["F2"] = "---"                                          #PART LABELID
        # sheet2["H2"] = 0                                          #PART_NAME    
        # sheet2["I2"] = 0                                              #LABOR
        # sheet2["I2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet2["J2"] = 0                                              #TOTAL_PRICE
        # sheet2["J2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet2["Q2"] = "---"

        set_column_width(sheet2)

    if len(ACCESSORY_PARTS_LIST) != 0:
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "PART_DIMENSIONS", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        for i in range(len(ACCESSORY_PARTS_LIST)):
            room_name = ACCESSORY_PARTS_LIST[i][0]         
            sku_num = ACCESSORY_PARTS_LIST[i][1]
            vendor_name = ACCESSORY_PARTS_LIST[i][2]           
            vendor_item = ACCESSORY_PARTS_LIST[i][3]
            part_id = ACCESSORY_PARTS_LIST[i][4] 
            part_name = ACCESSORY_PARTS_LIST[i][5]
            length = ACCESSORY_PARTS_LIST[i][6]
            width = ACCESSORY_PARTS_LIST[i][7]
            quantity = ACCESSORY_PARTS_LIST[i][8]
            franchise_price = ACCESSORY_PARTS_LIST[i][10]
            uom = ACCESSORY_PARTS_LIST[i][11]

            sheet3["A" + str((i + 1) + 1)] = room_name     
            sheet3["B" + str((i + 1) + 1)] = sku_num
            sheet3["C" + str((i + 1) + 1)] = vendor_name           
            sheet3["D" + str((i + 1) + 1)] = vendor_item
            sheet3["E" + str((i + 1) + 1)] = part_id           
            sheet3["F" + str((i + 1) + 1)] = part_name
            if "SF" in uom:
                sheet3["G" + str((i + 1) + 1)] = str(length) + " x " + str(width) + " x " + str(width) + " (" + str(get_square_footage(float(length), float(width))) + " SF)"
            elif "LF" in uom:
                sheet3["G" + str((i + 1) + 1)] = str(length) + " (" + str(get_linear_footage(float(length))) + " LF)"
            else:
                sheet3["G" + str((i + 1) + 1)] = 0
            sheet3["H" + str((i + 1) + 1)] = quantity
            sheet3["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            if "SF" in uom:
                sheet3["I" + str((i + 1) + 1)] = float(franchise_price) * get_square_footage(float(length), float(width))
            elif "LF" in uom:
                sheet3["I" + str((i + 1) + 1)] = float(franchise_price) * get_linear_footage(float(length))
            else:
                sheet3["I" + str((i + 1) + 1)] = float(franchise_price)
            
            sheet3["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            if "SF" in uom:
                sheet3["J" + str((i + 1) + 1)] = (float(franchise_price) * int(quantity)) * get_square_footage(float(length), float(width))
            elif "LF" in uom:
                sheet3["J" + str((i + 1) + 1)] = (float(franchise_price) * int(quantity)) * get_linear_footage(float(length))
            else:
                sheet3["J" + str((i + 1) + 1)] = float(franchise_price) * int(quantity)

        row_max = sheet3.max_row
        sheet3["K" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet3["K" + str(row_max + 3)] = "TOTAL"
        sheet3["K" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet3["K" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet3)
    else:
        print("Accessory Parts List Empty")
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "PART_DIMENSIONS", "QUANTITY", "PART_PRICE", "TOTAL_PRICE"])

        # sheet3["A2"] = "---"                                          #ROOM_NAME
        # sheet3["B2"] = "---"                                          #WALL_NAME
        # sheet3["C2"] = "---"                                          #SKU_NUMBER 
        # sheet3["D2"] = "---"                                          #VENDOR_NAME
        # sheet3["E2"] = "---"                                          #VENDOR_ITEM
        # sheet3["F2"] = "---"                                          #PART LABELID
        # sheet3["G2"] = "---"
        # sheet3["H2"] = 0                                          #PART_NAME    
        # sheet3["I2"] = 0                                              #LABOR
        # sheet3["I2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet3["J2"] = 0                                              #TOTAL_PRICE
        # sheet3["J2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet3["Q2"] = "---"

        set_column_width(sheet3)

    if len(UPGRADED_PANEL_PARTS_LIST) != 0:
        sheet4 = wb.create_sheet()
        sheet4.title = "Upgraded Panels"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PAINT_STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        for i in range(len(UPGRADED_PANEL_PARTS_LIST)):
            room_name = UPGRADED_PANEL_PARTS_LIST[i][0]
            sku_num = UPGRADED_PANEL_PARTS_LIST[i][1]
            vendor_name = UPGRADED_PANEL_PARTS_LIST[i][2]
            vendor_item = UPGRADED_PANEL_PARTS_LIST[i][3]
            part_id = UPGRADED_PANEL_PARTS_LIST[i][4]
            paint_stain = UPGRADED_PANEL_PARTS_LIST[i][5]
            part_name = UPGRADED_PANEL_PARTS_LIST[i][6]
            quantity = UPGRADED_PANEL_PARTS_LIST[i][7]
            length = UPGRADED_PANEL_PARTS_LIST[i][8]
            width = UPGRADED_PANEL_PARTS_LIST[i][9]    
            thickness = UPGRADED_PANEL_PARTS_LIST[i][10]
            franchise_price = UPGRADED_PANEL_PARTS_LIST[i][14]
            f_labor = UPGRADED_PANEL_PARTS_LIST[i][13]

            sheet4["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet4["B" + str((i + 1) + 1)] = sku_num                                        #SKU_NUMBER 
            sheet4["C" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet4["D" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet4["E" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            sheet4["F" + str((i + 1) + 1)] = paint_stain                                          #PAINT_STAIN_COLOR
            sheet4["G" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet4["H" + str((i + 1) + 1)] = quantity                                          #QUANTITY
            sheet4["I" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS          
            sheet4["J" + str((i + 1) + 1)] = thickness                                         #THICKNESS

            # if 'Glass' in part_name and 'Glass Shelf' not in part_name:
            #     if UPGRADED_PANEL_PARTS_LIST[i][15] is not None:
            #         if UPGRADED_PANEL_PARTS_LIST[i][15] == 'Yes':
            #             sheet4["K" + str((i + 1) + 1)] = get_square_footage(round((float(length) - 6.75)/2, 2),round(float(width) - 4.75, 2)) * int(quantity)
            #         else:
            #             sheet4["K" + str((i + 1) + 1)] = get_square_footage(round(float(length) - 4.75, 2),round(float(width) - 4.75, 2)) * int(quantity)        
            # else:
            sheet4["K" + str((i + 1) + 1)] = get_square_footage(float(length),float(width)) * int(quantity)
            
            sheet4["L" + str((i + 1) + 1)] = "---"
            sheet4["M" + str((i + 1) + 1)] = float(franchise_price)                                  #PART_PRICE
            sheet4["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["N" + str((i + 1) + 1)] = float(f_labor)                                  #LABOR
            sheet4["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["O" + str((i + 1) + 1)] = float(franchise_price)  #TOTAL_PRICE
            sheet4["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet4.max_row
        sheet4["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet4["P" + str(row_max + 3)] = "TOTAL"
        sheet4["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet4["P" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet4.column_dimensions['N'].hidden = True
        set_column_width(sheet4)
    else:
        print("Upgraded Panels Parts List Empty")
        sheet4 = wb.create_sheet()
        sheet4.title = "Upgraded Panels"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PAINT_STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        # sheet4["A2"] = "---"                                          #ROOM_NAME
        # sheet4["B2"] = "---"                                          #SKU_NUMBER 
        # sheet4["C2"] = "---"                                          #VENDOR_NAME
        # sheet4["D2"] = "---"                                          #VENDOR_ITEM
        # sheet4["E2"] = "---"                                          #PART LABELID
        # sheet4["F2"] = "---"                                          #MATERIAL
        # sheet4["G2"] = "---"                                          #PART_NAME    
        # sheet4["H2"] = 0                                              #QUANTITY
        # sheet4["I2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet4["J2"] = "---"                                          #THICKNESS
        # sheet4["K2"] = "---"
        # sheet4["L2"] = "---"
        # sheet4["M2"] = 0                                              #PART_PRICE
        # sheet4["M2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet4["N2"] = 0                                              #LABOR
        # sheet4["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet4["O2"] = 0                                              #TOTAL_PRICE
        # sheet4["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        sheet4.column_dimensions['N'].hidden = True
        set_column_width(sheet4)

    if len(GLASS_PARTS_LIST) != 0:
        sheet5 = wb.create_sheet()
        sheet5.title = "Glass"
        sheet5.append([
            "ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID",
            "MATERIAL", "PART_NAME", "QUANTITY", "PART_DIMENSIONS", "THICKNESS",
            "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        for i in range(len(GLASS_PARTS_LIST)):
            room_name = GLASS_PARTS_LIST[i][0]
            sku_num = GLASS_PARTS_LIST[i][1]
            vendor_name = GLASS_PARTS_LIST[i][2]
            vendor_item = GLASS_PARTS_LIST[i][3]
            part_id = GLASS_PARTS_LIST[i][4]
            material = GLASS_PARTS_LIST[i][5]
            part_name = GLASS_PARTS_LIST[i][6]
            quantity = GLASS_PARTS_LIST[i][7]
            length = GLASS_PARTS_LIST[i][8]
            width = GLASS_PARTS_LIST[i][9]
            thickness = GLASS_PARTS_LIST[i][10]
            franchise_price = GLASS_PARTS_LIST[i][14]
            f_labor = GLASS_PARTS_LIST[i][13]

            is_glass = 'Glass' in part_name and 'Glass Shelf' not in part_name
            if is_glass and GLASS_PARTS_LIST[i][15] == 'Yes':
                quantity = int(quantity) * 2

            if sku_num == "GL-0000010":
                thickness = 0.125
                quantity = int(quantity) * 2

            sheet5["A" + str((i + 1) + 1)] = room_name        # ROOM_NAME
            sheet5["B" + str((i + 1) + 1)] = sku_num          # SKU_NUMBER
            sheet5["C" + str((i + 1) + 1)] = vendor_name      # VENDOR_NAME
            sheet5["D" + str((i + 1) + 1)] = vendor_item      # VENDOR_ITEM
            sheet5["E" + str((i + 1) + 1)] = part_id          # PART LABELID
            sheet5["F" + str((i + 1) + 1)] = material         # MATERIAL
            sheet5["G" + str((i + 1) + 1)] = part_name        # PART_NAME
            sheet5["H" + str((i + 1) + 1)] = int(quantity)    # QUANTITY

            # Normalize indices for readability
            row_idx = i + 2  # (i + 1) + 1
            col_i = "I" + str(row_idx)
            col_j = "J" + str(row_idx)
            col_k = "K" + str(row_idx)

            # Check for glass part
            is_glass_shelf = 'Glass Shelf' in part_name
            is_glass = 'Glass' in part_name
            square_footage = None

            if is_glass_shelf:
                sheet5[col_i] = f"{length} x {width}"  # PART_DIMENSIONS
            elif is_glass:
                style_name = GLASS_PARTS_LIST[i][16]

                # If door has center rail
                if is_glass and GLASS_PARTS_LIST[i][15] == 'Yes':
                    adjusted_length, adjusted_width = get_glass_inset_size(length, width, True, style_name)
                    square_footage = get_square_footage(adjusted_length, adjusted_width)
                else:
                    adjusted_length, adjusted_width = get_glass_inset_size(length, width, False, style_name)
                    square_footage = None if not is_glass else get_square_footage(adjusted_length, adjusted_width)

                # Write to sheet
                sheet5[col_i] = f"{adjusted_length} x {adjusted_width}"  # PART_DIMENSIONS

            sheet5[col_j] = thickness  # THICKNESS
            if square_footage is not None:
                sheet5[col_k] = square_footage
            else:
                sheet5["K" + str((i + 1) + 1)] = get_square_footage(float(length),float(width))

            sheet5["L" + str((i + 1) + 1)] = get_linear_footage(float(length))
            sheet5["M" + str((i + 1) + 1)] = float(franchise_price)                                  #PART_PRICE
            sheet5["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["N" + str((i + 1) + 1)] = float(f_labor)                                  #LABOR
            sheet5["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["O" + str((i + 1) + 1)] = float(franchise_price) * int(quantity)
            sheet5["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet5.max_row
        sheet5["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet5["P" + str(row_max + 3)] = "TOTAL"
        sheet5["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet5["P" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet5.column_dimensions['N'].hidden = True
        set_column_width(sheet5)
    else:
        print("Glass Parts List Empty")
        sheet5 = wb.create_sheet()
        sheet5.title = "Glass"
        sheet5.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        # sheet5["A2"] = "---"                                          #ROOM_NAME
        # sheet5["B2"] = "---"                                          #SKU_NUMBER 
        # sheet5["C2"] = "---"                                          #VENDOR_NAME
        # sheet5["D2"] = "---"                                          #VENDOR_ITEM
        # sheet5["E2"] = "---"                                          #PART LABELID
        # sheet5["F2"] = "---"                                          #MATERIAL
        # sheet5["G2"] = "---"                                          #PART_NAME    
        # sheet5["H2"] = 0                                              #QUANTITY
        # sheet5["I2"] = "---" + " x " + "---"                          #PART_DIMENSIONS           
        # sheet5["J2"] = "---"                                          #THICKNESS
        # sheet5["K2"] = 0
        # sheet5["L2"] = 0
        # sheet5["M2"] = 0                                              #PART_PRICE
        # sheet5["M2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet5["N2"] = 0                                              #LABOR
        # sheet5["N2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet5["O2"] = 0                                              #TOTAL_PRICE
        # sheet5["O2"].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        sheet5.column_dimensions['N'].hidden = True
        set_column_width(sheet5)


    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        sheet6 = wb.create_sheet()
        sheet6.title = "Special Order"
        sheet6.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        for i in range(len(SPECIAL_ORDER_PARTS_LIST)):
            room_name = SPECIAL_ORDER_PARTS_LIST[i][0]
            sku_num = SPECIAL_ORDER_PARTS_LIST[i][1]
            vendor_name = SPECIAL_ORDER_PARTS_LIST[i][2]
            vendor_item = SPECIAL_ORDER_PARTS_LIST[i][3]
            part_id = SPECIAL_ORDER_PARTS_LIST[i][4]
            material = SPECIAL_ORDER_PARTS_LIST[i][5]
            part_name = SPECIAL_ORDER_PARTS_LIST[i][6]
            quantity = SPECIAL_ORDER_PARTS_LIST[i][7]
            length = SPECIAL_ORDER_PARTS_LIST[i][8]
            width = SPECIAL_ORDER_PARTS_LIST[i][9]    
            thickness = SPECIAL_ORDER_PARTS_LIST[i][10]
            franchise_price = SPECIAL_ORDER_PARTS_LIST[i][14]
            f_labor = SPECIAL_ORDER_PARTS_LIST[i][13]

            sheet6["A" + str((i + 1) + 1)] = room_name                                          #ROOM_NAME
            sheet6["B" + str((i + 1) + 1)] = sku_num                                          #SKU_NUMBER 
            sheet6["C" + str((i + 1) + 1)] = vendor_name                                          #VENDOR_NAME
            sheet6["D" + str((i + 1) + 1)] = vendor_item                                          #VENDOR_ITEM
            sheet6["E" + str((i + 1) + 1)] = part_id                                          #PART LABELID
            sheet6["F" + str((i + 1) + 1)] = material                                          #MATERIAL
            sheet6["G" + str((i + 1) + 1)] = part_name                                          #PART_NAME    
            sheet6["H" + str((i + 1) + 1)] = int(quantity)                                          #QUANTITY
            if 'Glass' in part_name and 'Glass Shelf' not in part_name:
                if SPECIAL_ORDER_PARTS_LIST[i][15] is not None:
                    if SPECIAL_ORDER_PARTS_LIST[i][15] == 'Yes':
                        sheet6["I" + str((i + 1) + 1)] = str(round((float(length) - 6.75)/2, 2)) + " x " + str(round(float(width) - 4.75, 2))
                    else:
                        sheet6["I" + str((i + 1) + 1)] = str(round(float(length) - 4.75, 2)) + " x " + str(round(float(width) - 4.75, 2))      #PART_DIMENSIONS
            else:
                sheet6["I" + str((i + 1) + 1)] = str(length) + " x " + str(width)      #PART_DIMENSIONS      
            sheet6["J" + str((i + 1) + 1)] = thickness                                         #THICKNESS
            if 'Glass' in part_name and 'Glass Shelf' not in part_name:
                if SPECIAL_ORDER_PARTS_LIST[i][15] is not None:
                    if SPECIAL_ORDER_PARTS_LIST[i][15] == 'Yes':
                        sheet6["K" + str((i + 1) + 1)] = get_square_footage(round((float(length) - 6.75)/2, 2),round(float(width) - 4.75, 2)) * int(quantity)
                    else:
                        sheet6["K" + str((i + 1) + 1)] = get_square_footage(round(float(length) - 4.75, 2),round(float(width) - 4.75, 2)) * int(quantity)
                        
            else:
                sheet6["K" + str((i + 1) + 1)] = get_square_footage(float(length),float(width)) * int(quantity)
            sheet6["L" + str((i + 1) + 1)] = get_linear_footage(float(length))
            sheet6["M" + str((i + 1) + 1)] = float(franchise_price)                                  #PART_PRICE
            sheet6["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet6["N" + str((i + 1) + 1)] = float(f_labor)                                  #LABOR
            sheet6["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet6["O" + str((i + 1) + 1)] = (float(franchise_price) + float(f_labor)) * int(quantity)
            sheet6["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        # row_max = sheet6.max_row
        # sheet6["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        # sheet6["P" + str(row_max + 3)] = "TOTAL"
        # sheet6["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        # sheet6["P" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        sheet6.column_dimensions['N'].hidden = True
        set_column_width(sheet6)
    else:
        print("Special Order Parts List Empty")
        sheet6 = wb.create_sheet()
        sheet6.title = "Special Order"
        sheet6.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "PART_PRICE", "LABOR", "TOTAL_PRICE"])

        sheet6["Q1"] = "Be sure to note ALL material changes and special order additions on your plans"
        sheet6["Q1"].fill = openpyxl.styles.PatternFill("solid", start_color="FFFF00")
        sheet6["Q1"].font = openpyxl.styles.Font(bold=True)
        sheet6["Q1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet6.column_dimensions['N'].hidden = True
        set_column_width(sheet6)

    # Save the Workbook
    try:
        wb.save(filename=parts_file)
        print("Franchise Pricing Parts List Generated")
        print("Creating Franchise Parts Summary...")
    except PermissionError:
        parts_file_name = os.path.basename(parts_file)
        message = "Cannot create parts list. Pricing Spreadsheet Open.\nPlease close the file:\n" + parts_file_name
        return bpy.ops.snap.message_box('INVOKE_DEFAULT', message=message, icon='ERROR')

    # Create Pivot table for Materials and Glass Square Footage on Summary Sheet
    df_materials = pandas.read_excel(
        parts_file,
        sheet_name='Materials'
    ).query(
        'SKU_NUMBER.str.contains("EB")==False and MATERIAL.str.contains("BBBB")==False',
        engine='python'
    )

    # Calculate and add column for total square footage
    df_materials['SQUARE_FT_TOTAL'] = df_materials['QUANTITY'] * df_materials['SQUARE_FT']

    # Aggregate square footage total
    sf_materials = pandas.pivot_table(df_materials, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT_TOTAL'], aggfunc=numpy.sum)

    # Rename SQUARE_FT_TOTAL column
    sf_materials = sf_materials.rename(columns={'SQUARE_FT_TOTAL': 'SQUARE_FT'})

    df_glass = pandas.read_excel(parts_file, sheet_name='Glass')
    sf_glass = pandas.pivot_table(df_glass, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT'], aggfunc=numpy.sum)

    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                                                right=openpyxl.styles.borders.Side(style='thin'), 
                                                top=openpyxl.styles.borders.Side(style='thin'), 
                                                bottom=openpyxl.styles.borders.Side(style='thin'))

    with pandas.ExcelWriter(parts_file, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        sf_row_start = 2

        if not sf_materials.empty and not sf_glass.empty:
            frames = [df_materials, df_glass]
            all_frames = pandas.concat(frames)
            all_data = pandas.pivot_table(all_frames, index=['ROOM_NAME', 'SKU_NUMBER', 'MATERIAL'], values=['SQUARE_FT'], aggfunc=numpy.sum)
            all_data.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=2)
            sf_row_start = (all_data.shape[0] + 1) + 4
            room_data = pandas.pivot_table(all_frames, index=['ROOM_NAME'], values=['SQUARE_FT'], aggfunc=numpy.sum)
            room_data.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=sf_row_start)
            sf_row_start = (all_data.shape[0] + 1) + (room_data.shape[0] + 1) + 4
            project_data = all_frames['SQUARE_FT'].sum()
            ws = writer.sheets["Franchise Pricing Summary"]
            ws["D" + str(sf_row_start+3)] = "Project Materials SF Total"
            ws["D" + str(sf_row_start+3)].font = openpyxl.styles.Font(bold=True)
            ws["D" + str(sf_row_start+3)].border = thin_border
            ws["E" + str(sf_row_start+3)] = project_data
            ws["D" + str(sf_row_start+4)] = "Project Materials Estimated Weight(lbs)"
            ws["D" + str(sf_row_start+4)].font = openpyxl.styles.Font(bold=True)
            ws["D" + str(sf_row_start+4)].border = thin_border
            ws["E" + str(sf_row_start+4)] = project_data * 3.5

        else:
            if not sf_materials.empty:
                # Calculate total square footage per room
                room_materials = df_materials.groupby('ROOM_NAME').apply(calc_room_total_sq_ft).reset_index(name='SQUARE_FT')

                sf_materials.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=2)
                sf_row_start = (sf_materials.shape[0] + 1) + 4

                # Write without the default index column
                room_materials.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=sf_row_start, index=False)

                # Apply borders to room names column
                for i in range(len(room_materials) + 1):
                    ws = writer.sheets["Franchise Pricing Summary"]
                    ws["D" + str(sf_row_start + i + 1)].border = thin_border

                sf_row_start = (sf_materials.shape[0] + 1) + (room_materials.shape[0] + 1) + 4

                # Calculate total square footage for project
                total_sq_ft = (df_materials['QUANTITY'] * df_materials['SQUARE_FT']).sum()

                ws = writer.sheets["Franchise Pricing Summary"]
                ws["D" + str(sf_row_start+3)] = "Project Material SF Total"
                ws["D" + str(sf_row_start+3)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start+3)].border = thin_border
                ws["E" + str(sf_row_start+3)] = total_sq_ft
                ws["D" + str(sf_row_start+4)] = "Project Material Estimated Weight(lbs)"
                ws["D" + str(sf_row_start+4)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start+4)].border = thin_border
                ws["E" + str(sf_row_start+4)] = total_sq_ft * 3.5

            if not sf_glass.empty:
                sf_glass.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=2)
                sf_row_start = (sf_glass.shape[0] + 1) + 3
                room_glass = pandas.pivot_table(df_glass, index=['ROOM_NAME'], values=['SQUARE_FT'], aggfunc=numpy.sum)
                room_glass.to_excel(writer, 'Franchise Pricing Summary', startcol=3, startrow=sf_row_start)
                sf_row_start = (sf_glass.shape[0] + 1) + (room_glass.shape[0] + 1) + 4
                total_sq_ft = (df_glass['QUANTITY'] * df_glass['SQUARE_FT']).sum()
                ws = writer.sheets["Franchise Pricing Summary"]
                ws["D" + str(sf_row_start+3)] = "Project Glass SF Total"
                ws["D" + str(sf_row_start+3)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start+3)].border = thin_border
                ws["E" + str(sf_row_start+3)] = total_sq_ft
                ws["D" + str(sf_row_start+4)] = "Project Glass Estimated Weight(lbs)"
                ws["D" + str(sf_row_start+4)].font = openpyxl.styles.Font(bold=True)
                ws["D" + str(sf_row_start+4)].border = thin_border
                ws["E" + str(sf_row_start+4)] = total_sq_ft * 3.5

        set_column_width(writer.sheets['Franchise Pricing Summary'])

    wb.save(filename=parts_file)
    generate_parts_summary(parts_file, sheet1, sheet2, sheet3, sheet4, sheet5, sheet6)


def get_labor_costs(part_name):
    if part_name is not None:
        part_name = part_name.upper()
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            DisplayName,\
            FranchisePrice,\
            RetailPrice\
        FROM\
            {CCItems}\
        WHERE\
            ProductType == 'LBR' AND\
            DisplayName LIKE 'LABOR - {}'\
        ;".format(part_name, CCItems="CCItems_" + sn_utils.get_franchise_location())
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        # print("No Labor Price Found for {}".format(part_name))
    for row in rows:
        sku = row[0]
        display_name = row[1]
        franchise_price = float(row[2])
        retail_price = float(row[3])
        # print("SKU: {}  DisplayName: {}  Part Name: {}  RetailPrice: {}".format(sku, display_name, part_name, retail_price))

    return retail_price, franchise_price


def get_eb_measurements(eb_orientation, length, width):
    if length > width:
        long_side = length
        short_side = width
    else:
        long_side = width
        short_side = length

    if eb_orientation == 'S1' or eb_orientation == 'S2':
        return short_side
    if eb_orientation == 'L1' or eb_orientation == 'L2':
        return long_side
    else:
        return long_side


def get_price_by_sku(sku_num):
    sku_num = str(sku_num)
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            DisplayName,\
            RetailPrice,\
            FranchisePrice,\
            UOM,\
            VendorName,\
            VendorItemNum\
        FROM\
            {CCItems}\
        WHERE\
            SKU = '{SKU}'\
        ;".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), SKU=sku_num)
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        uom = ''
        display_name = ''
        vendor_name = ''
        vendor_item = 0
        print("No Prices Returned for SKU:  " + sku_num)
    for row in rows:
        display_name = row[1]
        retail_price = float(row[2])
        franchise_price = float(row[3])
        uom = row[4]
        vendor_name = row[5]
        vendor_item = row[6]

    return retail_price, franchise_price


def get_glass_sku(glass_color):
    glass_thickness = 0.25
    if 'Frosted' in glass_color or 'Smoked' in glass_color:
        glass_thickness = 0.13
    if glass_color == 'Clear':
        glass_color = 'Clear Annealed'
    rows = sn_db.query_db(
        "SELECT\
            SKU\
        FROM\
            {CCItems}\
        WHERE\
            ProductType = 'GL' AND\
            Thickness == '{GlassThickness}' AND\
            DisplayName LIKE '%{DisplayName}%';\
        ".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), DisplayName=glass_color, GlassThickness=glass_thickness)
    )
    if len(rows) == 0:
        sku = 'SO-0000001'
        print("No Pricing SKU Returned for Material Name:  " + glass_color)
        print("Special Order SKU given to:  " + glass_color)
    for row in rows:
        sku = row[0]
    return sku


def get_glass_thickness_by_sku(glass_sku):
    glass_thickness = 0

    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            THICKNESS\
        FROM\
            {CCItems}\
        WHERE\
            SKU == '{Glass_SKU}';\
        ".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), Glass_SKU=glass_sku)
    )

    if len(rows) == 0:
        print("Could not get glass thickness for SKU: " + glass_sku)

    for row in rows:
        glass_thickness = row[1]

    return glass_thickness


def get_mat_sku(mat_name):
    rows = sn_db.query_db(
        "SELECT\
            SKU\
        FROM\
            {CCItems}\
        WHERE\
            DisplayName LIKE '%{DisplayName}%';\
        ".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), DisplayName=mat_name)
    )
    if len(rows) == 0:
        sku = 'SO-0000001'
        print("No Pricing SKU Returned for Material Name:  " + mat_name)
        print("Special Order SKU given to:  " + mat_name)
    for row in rows:
        sku = row[0]
    return sku


def get_paint_stain_mat_sku(mat_name):
    name, product_type = mat_name.split('-')

    if "Other/Custom" in name:
        name = "Custom Match"

    rows = sn_db.query_db(
        "SELECT\
            SKU\
        FROM\
            {CCItems}\
        WHERE\
            ProductType IN ('{Product_Type}') AND DisplayName LIKE '%{DisplayName}%';\
        ".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), Product_Type=product_type.strip(), DisplayName=name.strip())
    )
    if len(rows) == 0:
        sku = 'SO-0000001'
        print("No Pricing SKU Returned for Material Name:  " + mat_name)
        print("Special Order SKU given to:  " + mat_name)
    for row in rows:
        sku = row[0]

    return sku

def get_mat_display_name(mat_sku):
    rows = sn_db.query_db(
        "SELECT\
            DisplayName\
        FROM\
            {CCItems}\
        WHERE\
            SKU LIKE '%{mat_sku}%';\
        ".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), mat_sku=mat_sku)
    )
    if len(rows) == 0:
        display_name = 'Unknown'
        print("No display name found for the following sku:  " + mat_sku)
        print("Unknown display name given to:  " + mat_sku)
    for row in rows:
        display_name = row[0]
    return display_name


def get_pricing_info(sku_num, qty, length_inches=0.0, width_inches=0.0, style_name=None, is_glaze=False, glaze_style=None, glaze_color=None, part_name='', eb_orientation=None, center_rail=None, upgrade_color=None, BoxAssembled=None):
    length_inches = float(length_inches)
    width_inches = float(width_inches)
    upgraded_panel_pricing_file = os.path.join(sn_paths.ROOT_DIR, "db_init", "Upgraded_Panel_Pricing.xlsx")
    style_type = ''
    glaze_price = 0.0
    r_labor_price = 0
    f_labor_price = 0
    eb_length = 0
    paint_stain_price = [0,0]
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            Name,\
            RetailPrice,\
            FranchisePrice,\
            UOM,\
            VendorName,\
            VendorItemNum\
        FROM\
            {CCItems}\
        WHERE\
            SKU = '{SKU}'\
        ;".format(CCItems="CCItems_" + sn_utils.get_franchise_location(), SKU=sku_num)
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        uom = ''
        name = ''
        vendor_name = ''
        vendor_item = 0
        print("No Prices Returned for SKU:  " + sku_num)
    for row in rows:
        name = row[1]

        if type(row[2]) == str:
            retail_price = float(row[2].replace(',', ''))
        else:
            retail_price = float(row[2])

        franchise_price = float(row[3])
        uom = row[4]
        vendor_name = row[5]
        vendor_item = row[6]

    price_check(sku_num, franchise_price, retail_price)

    if part_name is not None:
        # Get Labor Costs
        if sku_num[:2] in material_types and not sku_num[:2] in 'EB':
            labor = get_labor_costs(part_name)
            drawer_bottom_part_names = [
                "Drawer Bottom",
                "DrwrBox Bottom - Mel",
                "DrwrBox Bttm DT - BB",
                "DrwrBox Inset Bttm - Mel",
                "DrwrBox Inset Bttm - BB"
            ]

            part_name = part_name.strip()

            if part_name in drawer_bottom_part_names:
                # For XMLs generated pre-280. Drawer box bottom part names updated to match COS with v280 (#1893)
                if "Drawer Bottom" in part_name:
                    if 'BB' in sku_num[:2]:
                        labor = get_labor_costs('BB DRAWER INSET BOTTOM')
                    else:
                        labor = get_labor_costs('MEL DRAWER CAP BOTTOM')
                # 280+ and COS drawer bottoms
                elif part_name == "DrwrBox Bottom - Mel":
                    labor = get_labor_costs('MEL DRAWER CAP BOTTOM')
                elif part_name == "DrwrBox Bttm DT - BB":
                    labor = get_labor_costs('BB DRAWER CAP BTTM')
                elif part_name == "DrwrBox Inset Bttm - Mel":
                    labor = get_labor_costs('MEL DRAWER INSET BTTM')
                elif part_name == "DrwrBox Inset Bttm - BB":
                    labor = get_labor_costs('BB DRAWER INSET BOTTOM')
                if BoxAssembled and BoxAssembled != 'False':
                    assembled_labor = get_price_by_sku('LB-0000022')
                    labor_list = list(labor)
                    labor_list[0] = labor[0] + assembled_labor[0]
                    labor_list[1] = labor[1] + assembled_labor[1]
                    labor = tuple(labor_list)

            R_LABOR_PRICES.append(labor[0] * int(qty))
            F_LABOR_PRICES.append(labor[1] * int(qty))
            r_labor_price = labor[0]
            f_labor_price = labor[1]

    if 'SF' in uom and not sku_num[:2] in hardware_types and not sku_num[:2] in accessory_types and 'PL' not in sku_num[:2] and 'SN' not in sku_num[:2]:
        # Check if Melamine material contains a style_name that makes it a 5PC Melamine Door
        if style_name is not None and 'VN' not in sku_num[:2] and 'WD' not in sku_num[:2] and 'GL' not in sku_num[:2] and 'Slab Door' not in style_name:
            if 'Traviso' in style_name:
                door_labor_price = get_price_by_sku('LB-0000013')
                R_LABOR_PRICES.append(door_labor_price[0] * int(qty))
                F_LABOR_PRICES.append(door_labor_price[1] * int(qty))
                r_labor_price += door_labor_price[0]
                f_labor_price += door_labor_price[1]
            else:
                door_labor_price = (0, 0)
            if 'Glass' in style_name:
                if center_rail == 'Yes':
                    sf_door_glass = get_square_footage(((length_inches-6.75)/2), width_inches-4.75)
                else:
                    sf_door_glass = get_square_footage(length_inches-4.75, width_inches-4.75)
                retail_glass_door_labor = 0
                franchise_glass_door_labor = 0
                # Melamine and Traviso Glass Door Labor Charges
                if 'Traviso' in style_name:
                    glass_door_labor_charge = get_price_by_sku('LB-0000020')
                    R_LABOR_PRICES.append(glass_door_labor_charge[0] * int(qty))
                    F_LABOR_PRICES.append(glass_door_labor_charge[1] * int(qty))
                    retail_glass_door_labor = glass_door_labor_charge[0]
                    franchise_glass_door_labor = glass_door_labor_charge[1]
                elif "Melamine Door" in style_name:
                    glass_door_labor_charge = get_price_by_sku('LB-0000021')
                    R_LABOR_PRICES.append(glass_door_labor_charge[0] * int(qty))
                    F_LABOR_PRICES.append(glass_door_labor_charge[1] * int(qty))
                    retail_glass_door_labor = glass_door_labor_charge[0]
                    franchise_glass_door_labor = glass_door_labor_charge[1]
                retail_price = ((((get_square_footage(length_inches, width_inches) - sf_door_glass) * retail_price) + door_labor_price[0] + retail_glass_door_labor) * int(qty))
                franchise_price = ((((get_square_footage(length_inches, width_inches) - sf_door_glass) * franchise_price) + door_labor_price[1] + franchise_glass_door_labor) * int(qty))
                R_UPGRADED_PANEL_PRICES.append(retail_price)
                F_UPGRADED_PANEL_PRICES.append(franchise_price)
                r_labor_price += glass_door_labor_charge[0]
                f_labor_price += glass_door_labor_charge[1]
            else:
                retail_price = (((get_square_footage(length_inches, width_inches) * retail_price) + door_labor_price[0]) * int(qty))
                franchise_price = (((get_square_footage(length_inches, width_inches) * franchise_price) + door_labor_price[1]) * int(qty))
                R_UPGRADED_PANEL_PRICES.append(retail_price)
                F_UPGRADED_PANEL_PRICES.append(franchise_price)
        else:
            is_vn_sku = 'VN' in sku_num[:2]
            is_wd_sku = 'WD' in sku_num[:2]
            is_pbi_sku = 'pbi' in name.lower()

            if is_vn_sku or is_wd_sku and not is_pbi_sku:
                if upgrade_color is not None:
                    paint_stain_price = get_price_by_sku(get_paint_stain_mat_sku(upgrade_color))
                paint_stain_labor = get_price_by_sku('LB-0000014')
                sf_part = get_square_footage(length_inches, width_inches)
                r_stain_total = (paint_stain_price[0] * sf_part) * 2
                f_stain_total = (paint_stain_price[1] * sf_part) * 2
                r_part_total = retail_price * sf_part
                f_part_total = franchise_price * sf_part
                retail_price = r_part_total + r_stain_total + paint_stain_labor[0]
                franchise_price = f_part_total + f_stain_total + paint_stain_labor[1]
                R_MATERIAL_PRICES.append(retail_price * int(qty))
                R_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))
                F_MATERIAL_PRICES.append(franchise_price * int(qty))
                F_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))
            elif 'GL' in sku_num[:2]:
                if 'Glass Shelf' not in part_name:
                    if center_rail == 'Yes':
                        sf_glass = get_square_footage(((length_inches-6.75)/2), width_inches-4.75)
                    else:
                        sf_glass = get_square_footage(length_inches-4.75, width_inches-4.75)
                    retail_price = (retail_price * sf_glass)
                    franchise_price = (franchise_price * sf_glass)
                    R_GLASS_PRICES.append(retail_price * int(qty))
                    F_GLASS_PRICES.append(franchise_price * int(qty))
                else:
                    retail_price = (((get_square_footage(length_inches, width_inches)) * retail_price) + r_labor_price)
                    franchise_price = (((get_square_footage(length_inches, width_inches)) * franchise_price) + f_labor_price)
                    R_GLASS_PRICES.append(retail_price * int(qty))
                    F_GLASS_PRICES.append(franchise_price * int(qty))
            else:
                # Retail price for flat crown = $17/LF
                if part_name == "Flat Crown":
                    price = (((get_linear_footage(length_inches) * 17.0)) * int(qty))
                    R_MATERIAL_PRICES.append(price)
                    retail_price = 17.0
                else:
                    # Melamine Door Labor Charges
                    if "Door" in part_name:
                        add_mel_door_labor = not style_name or "Slab Door" in style_name
                        if add_mel_door_labor:
                            r_labor_price, f_labor_price = get_price_by_sku('LB-0000012')

                    R_MATERIAL_PRICES.append((((get_square_footage(length_inches, width_inches)) * retail_price) + r_labor_price) * int(qty))
                    R_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))

                F_MATERIAL_PRICES.append((((get_square_footage(length_inches, width_inches)) * franchise_price) + f_labor_price) * int(qty))
                F_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))

    if 'LF' in uom and not sku_num[:2] in hardware_types and not sku_num[:2] in accessory_types:
        if eb_orientation is not None:
            eb_length = get_eb_measurements(eb_orientation, length_inches, width_inches)
        R_MATERIAL_PRICES.append((get_linear_footage(eb_length) * retail_price) * int(qty))
        R_MATERIAL_LINEAR_FOOTAGE.append(get_linear_footage(eb_length))
        F_MATERIAL_PRICES.append((get_linear_footage(eb_length) * franchise_price) * int(qty))
        F_MATERIAL_LINEAR_FOOTAGE.append(get_linear_footage(eb_length))
    if 'PL' in sku_num[:2] or 'SN' in sku_num[:2]:
        r_stain_total = 0
        r_door_total = 0
        f_stain_total = 0
        f_door_total = 0
        veneer_price = get_price_by_sku('VN-0000005')
        three_quarter_mdf_price = get_price_by_sku('WD-0000010')
        quarter_mdf_price = get_price_by_sku('WD-0000007')
        upgraded_panel_labor_price = get_price_by_sku('LB-0000014')
        primer_price = 0
        sealer_price = get_price_by_sku('CH-0000001')
        topcoat_price = get_price_by_sku('CH-0000002')
        catalyst_price = get_price_by_sku('CH-0000003')
        jacketboard_price = get_price_by_sku('WD-0000002')
        care_price = get_price_by_sku('CH-0000004')
        glaze_labor = 0
    
        if style_name is not None:
            # try:
            #     import pandas
            #     import numpy
            # except ModuleNotFoundError:
            #     python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
            #     sys.path.append(python_lib_path)

            upgraded_panel_data_frame = pandas.read_excel(upgraded_panel_pricing_file, header=None, skiprows=1, usecols=[0, 1, 2, 3, 4, 5], names=['STYLE', 'TYPE', 'PANEL', 'SKU', 'AP_SKU', 'GL_NUM'])

            if 'door'.upper() in style_name.upper():
                style_type = 'Door'
            else:
                style_type = 'Drawer'

            for index, data_row in upgraded_panel_data_frame.iterrows():
                if data_row['STYLE'] in style_name and style_type == data_row['TYPE']:
                    panel = data_row['PANEL']
                    molding_sku = data_row['SKU']
                    ap_molding_sku = data_row['AP_SKU']
                    glazing_lines = data_row['GL_NUM']
                    lf_molding = round(((length_inches / 12) + (width_inches / 12)) * 2, 2)
                    lf_glaze = round(((length_inches / 12) + (width_inches / 12)) * 2, 2) * glazing_lines
                    molding_price = get_price_by_sku(molding_sku)
                    ap_molding_price = get_price_by_sku(ap_molding_sku)
                    sf_center_panel = get_square_footage(length_inches-4.75, width_inches-4.75)
                    if 'Glass' in style_name:
                        if center_rail == 'Yes':
                            sf_door_glass = get_square_footage(((length_inches-6.75)/2), width_inches-4.75)
                        else:
                            sf_door_glass = get_square_footage(length_inches-4.75, width_inches-4.75)
                        sf_door = get_square_footage(length_inches, width_inches) - sf_door_glass
                    else:
                        sf_door = get_square_footage(length_inches, width_inches)
                    
                    r_labor_price = upgraded_panel_labor_price[0]
                    f_labor_price = upgraded_panel_labor_price[1]
                    
                    if 'SN' in sku_num[:2]:
                        if is_glaze:
                            glaze_labor = get_labor_costs(glaze_style)
                            r_stain_total = (retail_price * sf_door) + (sealer_price[0] * sf_door) + (lf_glaze * glaze_labor[0]) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (franchise_price * sf_door) + (sealer_price[1] * sf_door) + (lf_glaze * glaze_labor[1]) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        else:
                            r_stain_total = (retail_price * sf_door) + (sealer_price[0] * sf_door) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (franchise_price * sf_door) + (sealer_price[1] * sf_door) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        if 'Raised' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * jacketboard_price[0]) + upgraded_panel_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * jacketboard_price[1]) + upgraded_panel_labor_price[1]
                        if 'Recessed' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * veneer_price[0]) + upgraded_panel_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * veneer_price[1]) + upgraded_panel_labor_price[1]

                    if 'PL' in sku_num[:2]:
                        if 'PL-0000004' or 'PL-0000005' or 'PL-0000009' in sku_num:
                            primer_price = get_price_by_sku('CH-0000005')
                        else:
                            primer_price = get_price_by_sku('CH-0000006')
                        if is_glaze:
                            glaze_labor = get_labor_costs(glaze_style)
                            r_stain_total = (primer_price[0] * sf_door) + (retail_price * sf_door) + (sealer_price[0] * sf_door) + (lf_glaze * glaze_labor[0]) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (primer_price[1] * sf_door) + (franchise_price * sf_door) + (sealer_price[1] * sf_door) + (lf_glaze * glaze_labor[1]) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        else:   
                            r_stain_total = (primer_price[0] * sf_door) + ((retail_price + care_price[0]) * sf_door)
                            f_stain_total = (primer_price[1] * sf_door) + ((franchise_price + care_price[1]) * sf_door)
                        if 'Raised' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * three_quarter_mdf_price[0]) + upgraded_panel_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * three_quarter_mdf_price[1]) + upgraded_panel_labor_price[1]
                        else:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * quarter_mdf_price[0]) + upgraded_panel_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * quarter_mdf_price[1]) + upgraded_panel_labor_price[1]

                    R_LABOR_PRICES.append(upgraded_panel_labor_price[0] * int(qty))
                    F_LABOR_PRICES.append(upgraded_panel_labor_price[1] * int(qty))
                    R_UPGRADED_PANEL_PRICES.append((r_door_total + r_stain_total) * int(qty))
                    F_UPGRADED_PANEL_PRICES.append((f_door_total + f_stain_total) * int(qty))
                    retail_price = r_door_total + r_stain_total
                    franchise_price = f_door_total + f_stain_total
    if sku_num[:2] in hardware_types:
        R_HARDWARE_PRICES.append(retail_price * int(qty))
        F_HARDWARE_PRICES.append(franchise_price * int(qty))
    if sku_num[:2] in accessory_types:
        if "Wire Basket" in part_name:
            retail_price = 85.0
            franchise_price = 28.0
        if 'SF' in uom:
            R_ACCESSORY_PRICES.append(get_square_footage(length_inches, width_inches) * (retail_price * int(qty)))
            F_ACCESSORY_PRICES.append(get_square_footage(length_inches, width_inches) * (franchise_price * int(qty)))
        elif 'LF' in uom:
            R_ACCESSORY_PRICES.append(get_linear_footage(length_inches) * (retail_price * int(qty)))
            F_ACCESSORY_PRICES.append(get_linear_footage(length_inches) * (franchise_price * int(qty)))
        else:
            R_ACCESSORY_PRICES.append(retail_price * int(qty))
            F_ACCESSORY_PRICES.append(franchise_price * int(qty))
    if sku_num[:2] in special_order_types:
        R_SPECIAL_ORDER_PRICES.append((retail_price * int(qty)) + r_labor_price)
        F_SPECIAL_ORDER_PRICES.append((franchise_price * int(qty)) + f_labor_price)
    
    if ('SF' in uom and sku_num[:2] in material_types) or 'EB' in sku_num[:2]:  
        name = get_mat_display_name(sku_num)

    return str(retail_price), str(franchise_price), name, vendor_name, vendor_item, style_name, r_labor_price, f_labor_price, uom


def calculate_project_price(xml_file, cos_flag=False):
    tree = None
    root = None
    eb_orientation = None
    eb_counter = 1
    global COS_FLAG
    global PROJECT_NAME
    COS_FLAG = cos_flag
    xml_file_exists = False

    if COS_FLAG:
        global PROJECT_NAME
        PROJECT_NAME = 'COS_Pricing_Project'

    if os.path.exists(xml_file):
        xml_file_exists = True
        tree = ET.parse(xml_file)
        root = tree.getroot().find("Job")
    else:
        print("XML filepath does not exist:", xml_file)
        xml_file_exists = False
        return xml_file_exists

    set_job_info(root)

    mfg = root.find("Manufacturing")
                                
    #Collect and Output
    for item in root.findall("Item"):

        # for item_name in item.findall("Name"):

        for description in item.findall("Description"):
            DESCRIPTION = description.text

        for var in item.findall("Var"):
            if var.find("Name").text == "DrawingNum":
                ROOM_DRAWING_NUMBER = var.find("Value").text

        for part in item.findall("Part"):
            PART_LABEL_ID = part.attrib.get('LabelID')

            for quantity in part.findall("Quantity"):
                QUANTITY = quantity.text

            for part_name in part.findall("Name"):
                if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                    PART_NAME = "Cleat"
                elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                    PART_NAME = "Drawer Side"
                elif part_name.text == "Left Door" or part_name.text == "Right Door":
                    PART_NAME = "Door"
                else:
                    PART_NAME = part_name.text

                for width in part.findall("Width"):
                    WIDTH = width.text

                for length in part.findall("Length"):
                    LENGTH = length.text

                for thickness in part.findall("Thickness"):
                    THICKNESS = thickness.text

                for part_type in part.findall("Type"):
                    PART_TYPE = part_type.text

            for label in mfg.findall("Label"):
                label_id = label.attrib.get("ID", None)
                # part_id = part.attrib.get("PartID", None)
                dcname = None
                eb1 = None
                eb2 = None
                eb3 = None
                eb4 = None
                rod_length = None
                style_name = 'None'
                glaze_style = 'None'
                glaze_color = 'None'
                is_glaze = False
                glass_color = 'None'
                wall_name = None
                center_rail = 'None'
                drill_pattern = '-'
                pull_dim = "-"
                BoxAssembled = 'False'
                lenx = 0
                leny = 0
                upgrade_color = None
                if PART_LABEL_ID == label_id:
                    print(PART_NAME + " :: " + PART_LABEL_ID)
                    # Use the iterator to deal with repeated children and still in order
                    label_iterator = iter(label)
                    for _ in range(int(len(label) / 3)):
                        sku_value = None
                        NAME = next(label_iterator).text
                        TYPE = next(label_iterator).text
                        VALUE = next(label_iterator).text
                        if NAME == 'dcname':
                            if VALUE == "Top Cleat" or VALUE == "Bottom Cleat":
                                VALUE = "Cleat"
                            if VALUE == "Left Drawer Side" or VALUE == "Right Drawer Side":
                                VALUE = "Drawer Side"
                            if VALUE == "Left Door" or VALUE == "Right Door":
                                VALUE = "Door"
                            dcname = VALUE
                        if 'sku' in NAME:
                            sku_value = VALUE
                            if sku_value == 'Unknown':
                                sku_value = 'SO-0000001'
                            if NAME == 'sku':
                                if sku_value == 'None':
                                    sku_value = 'SO-0000001'
                        if NAME == 'wallname':
                            wall_name = VALUE
                        if NAME == 'lenx':
                            lenx = VALUE
                        if NAME == 'leny':
                            leny = VALUE
                        if NAME == 'upgradeoption':
                            upgrade_color = VALUE
                        if NAME == 'edgeband1':
                            eb1 = VALUE
                        if NAME == 'edgeband2':
                            eb2 = VALUE
                        if NAME == 'edgeband3':
                            eb3 = VALUE
                        if NAME == 'edgeband4':
                            eb4 = VALUE
                        if 'rod'.upper() in PART_NAME.upper():
                            if NAME == 'lenx':
                                rod_length = VALUE
                        if NAME == 'style':
                            style_name = VALUE
                        if NAME == 'glazecolor':
                            glaze_color = VALUE
                            if glaze_color != 'None':
                                is_glaze = True
                        if NAME == 'glazestyle':
                            glaze_style = VALUE
                            if glaze_style != 'None':
                                is_glaze = True
                        if NAME == 'glasscolor':
                            glass_color = VALUE
                        if NAME == 'has_center_rail':
                            center_rail = VALUE
                        if NAME == 'drill_pattern':
                            drill_pattern = VALUE
                        if NAME == 'pull_dim':
                            pull_dim = VALUE
                        if NAME == 'BoxAssembled':
                            BoxAssembled = VALUE
                    # for sku_value in label.findall("Value"):
                        if sku_value is not None:
                            if sku_value[:2] in material_types:
                                SKU_NUMBER = sku_value
                                if sku_value[:2] in 'EB':
                                    if eb1 is not None and eb_counter == 1:
                                        eb_orientation = eb1
                                        eb_counter = 2
                                        if 's' in eb_orientation or 'S' in eb_orientation:
                                            S_COUNT.append(eb_orientation)
                                        if 'l' in eb_orientation or 'L' in eb_orientation:
                                            L_COUNT.append(eb_orientation)
                                        print(SKU_NUMBER + " :: " + eb_orientation)
                                    if eb2 is not None and eb_counter == 2 or eb2 is not None and eb_counter == 1:
                                        eb_orientation = eb2
                                        eb_counter = 3
                                        if 's' in eb_orientation or 'S' in eb_orientation:
                                            S_COUNT.append(eb_orientation)
                                        if 'l' in eb_orientation or 'L' in eb_orientation:
                                            L_COUNT.append(eb_orientation)
                                        print(SKU_NUMBER + " :: " + eb_orientation)
                                    if eb3 is not None and eb_counter == 3 or eb3 is not None and eb_counter == 2:
                                        eb_orientation = eb3
                                        eb_counter = 4
                                        if 's' in eb_orientation or 'S' in eb_orientation:
                                            S_COUNT.append(eb_orientation)
                                        if 'l' in eb_orientation or 'L' in eb_orientation:
                                            L_COUNT.append(eb_orientation)
                                        print(SKU_NUMBER + " :: " + eb_orientation)
                                    if eb4 is not None and eb_counter == 4 or eb4 is not None and eb_counter == 3:
                                        eb_orientation = eb4
                                        eb_counter = 1
                                        if 's' in eb_orientation or 'S' in eb_orientation:
                                            S_COUNT.append(eb_orientation)
                                        if 'l' in eb_orientation or 'L' in eb_orientation:
                                            L_COUNT.append(eb_orientation)
                                        print(SKU_NUMBER + " :: " + eb_orientation)
                                else:
                                    print(SKU_NUMBER)
                                # Continue statement to skip over edgebanding for glass shelves
                                if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                    continue
                                # Continue statement to skip over edgebanding for Upgraded Panels doors
                                if style_name != 'None' and sku_value[:2] in 'EB' and 'Slab Door' not in style_name:
                                    continue
                                if style_name != 'None' and sku_value[:2] not in 'EB' and 'Slab Door' not in style_name and sku_value[:2] not in 'GL':
                                    # pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                    if is_glaze:
                                        if glaze_color != 'None' and glaze_style != 'None':
                                            if glass_color != 'None':
                                                PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                            else:
                                                PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                        else:
                                            if glass_color != 'None':
                                                PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + ")"
                                            else:
                                                PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + ")"
                                    else:
                                        if glass_color != 'None':
                                            PART_NAME = dcname + " (" + style_name + "(" + glass_color + "))"
                                        else:
                                            PART_NAME = dcname + " (" + style_name + ")"

                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                    if 'Slab Door' in style_name:
                                        MATERIAL_PARTS_LIST.append([
                                            DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2],
                                            PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8],
                                            eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled])
                                    else:
                                        UPGRADED_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                    #Added for user to price the Glass within the upgraded panel parts separately
                                    # if 'Glass' in PART_NAME or 'Glass' in style_name:
                                    #     # PART_NAME = style_name + " (" + glass_color + ")"
                                    #     PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                    #     if 'GL' not in sku_value:
                                    #         SKU_NUMBER = get_glass_sku(glass_color)
                                    #         pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                    #     GLASS_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                else:
                                    if 'Glass' in PART_NAME or 'Glass' in style_name:
                                        if 'Glass Shelf' in PART_NAME:
                                            PART_NAME = dcname + " (" + glass_color + ")"
                                        else:
                                            print("    Glass SKU: " + sku_value, "Glass Color: " + glass_color, "Style Name: " + style_name, dcname)
                                            PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                            GLASS_THICKNESS = get_glass_thickness_by_sku(sku_value)
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail, upgrade_color)
                                        GLASS_PARTS_LIST.append(
                                            [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME,
                                             QUANTITY, LENGTH, WIDTH, GLASS_THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1],
                                             center_rail, style_name])
                                    else:
                                        if upgrade_color is not None:
                                            PART_NAME = dcname + " (" + upgrade_color + ")"
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation, center_rail, upgrade_color)

                                        MATERIAL_PARTS_LIST.append([
                                            DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4],PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH,
                                            THICKNESS, pricing_info[0],pricing_info[6], pricing_info[8], eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled
                                        ])
                            if sku_value[:2] in hardware_types:
                                SKU_NUMBER = sku_value
                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                PART_NAME = pricing_info[2]
                                HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                            if sku_value[:2] in accessory_types:
                                SKU_NUMBER = sku_value
                                if 'rod'.upper() in PART_NAME.upper():
                                    LENGTH = rod_length
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                else:
                                    LENGTH = lenx
                                    WIDTH = leny
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                    PART_NAME = pricing_info[2][:-10]
                                else:
                                    PART_NAME = pricing_info[2]
                                ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, WIDTH, QUANTITY, pricing_info[0], pricing_info[1], pricing_info[8]])
                            if sku_value[:2] in special_order_types:
                                LENGTH = lenx
                                WIDTH = leny
                                if 'Glass' in PART_NAME or 'Glass' in style_name:
                                    PART_NAME = dcname + " (" + glass_color + ")"
                                SKU_NUMBER = sku_value
                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, None, center_rail)
                                SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                    EDGEBANDING.append([PART_LABEL_ID, len(S_COUNT), len(L_COUNT)])
                    S_COUNT.clear()
                    L_COUNT.clear()
            eb_counter = 1
        LENGTH = 0
        WIDTH = 0
        THICKNESS = 0

        for assembly in item.findall("Assembly"):

            for part in assembly.findall("Part"):
                DESCRIPTION = description.text
                PART_LABEL_ID = part.attrib.get('LabelID')

                for quantity in part.findall("Quantity"):
                    QUANTITY = quantity.text

                for part_name in part.findall("Name"):
                    if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                        PART_NAME = "Cleat"
                    elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                        PART_NAME = "Drawer Side"
                    elif part_name.text == "Left Door" or part_name.text == "Right Door":
                        PART_NAME = "Door"
                    else:
                        PART_NAME = part_name.text

                    if part_name.text is not None:
                        for width in part.findall("Width"):
                            WIDTH = width.text

                        for length in part.findall("Length"):
                            LENGTH = length.text

                        for thickness in part.findall("Thickness"):
                            THICKNESS = thickness.text

                        for part_type in part.findall("Type"):
                            PART_TYPE = part_type.text

                for label in mfg.findall("Label"):
                    label_id = label.attrib.get("ID", None)
                    # part_id = part.attrib.get("PartID", None)
                    dcname = None
                    eb1 = None
                    eb2 = None
                    eb3 = None
                    eb4 = None
                    style_name = 'None'
                    glaze_style = 'None'
                    glaze_color = 'None'
                    is_glaze = False
                    glass_color = 'None'
                    wall_name = None
                    center_rail = 'None'
                    lenx = 0
                    leny = 0
                    upgrade_color = None
                    drill_pattern = '-'
                    pull_dim = "-"
                    BoxAssembled = 'False'
                    
                    if PART_LABEL_ID == label_id:
                        print(PART_NAME + " :: " + PART_LABEL_ID)
                        # Use the iterator to deal with repeated children and still in order
                        label_iterator = iter(label)
                        for _ in range(int(len(label) / 3)):
                            sku_value = None
                            NAME = next(label_iterator).text
                            TYPE = next(label_iterator).text
                            VALUE = next(label_iterator).text
                            if NAME == 'dcname':
                                if VALUE == "Top Cleat" or VALUE == "Bottom Cleat":
                                    VALUE = "Cleat"
                                if VALUE == "Left Drawer Side" or VALUE == "Right Drawer Side":
                                    VALUE = "Drawer Side"
                                if VALUE == "Left Door" or VALUE == "Right Door":
                                    VALUE = "Door"
                                dcname = VALUE
                            if 'sku' in NAME:
                                sku_value = VALUE
                                if sku_value == 'Unknown':
                                    sku_value = 'SO-0000001'
                                if NAME == 'sku':
                                    if sku_value == 'None':
                                        sku_value = 'SO-0000001'
                            if NAME == 'wallname':
                                wall_name = VALUE
                            if NAME == 'lenx':
                                lenx = VALUE
                            if NAME == 'leny':
                                leny = VALUE
                            if NAME == 'upgradeoption':
                                upgrade_color = VALUE
                            if NAME == 'edgeband1':
                                eb1 = VALUE
                            if NAME == 'edgeband2':
                                eb2 = VALUE
                            if NAME == 'edgeband3':
                                eb3 = VALUE
                            if NAME == 'edgeband4':
                                eb4 = VALUE
                            if 'rod'.upper() in PART_NAME.upper():
                                if NAME == 'lenx':
                                    rod_length = VALUE
                            if NAME == 'style':
                                style_name = VALUE
                            if NAME == 'glazecolor':
                                glaze_color = VALUE
                                if glaze_color != 'None':
                                    is_glaze = True
                            if NAME == 'glazestyle':
                                glaze_style = VALUE
                                if glaze_style != 'None':
                                    is_glaze = True
                            if NAME == 'glasscolor':
                                glass_color = VALUE
                            if NAME == 'has_center_rail':
                                center_rail = VALUE
                            if NAME == 'drill_pattern':
                                drill_pattern = VALUE
                            if NAME == 'pull_dim':
                                pull_dim = VALUE
                            if NAME == 'BoxAssembled':
                                BoxAssembled = VALUE
                        # for sku_value in label.findall("Value"):
                            if sku_value is not None:
                                if sku_value[:2] in material_types:
                                    SKU_NUMBER = sku_value
                                    if sku_value[:2] in 'EB':
                                        if eb1 is not None and eb_counter == 1:
                                            eb_orientation = eb1
                                            eb_counter = 2
                                            if 's' in eb_orientation or 'S' in eb_orientation:
                                                S_COUNT.append(eb_orientation)
                                            if 'l' in eb_orientation or 'L' in eb_orientation:
                                                L_COUNT.append(eb_orientation)
                                            print(SKU_NUMBER + " :: " + eb_orientation)
                                        if eb2 is not None and eb_counter == 2 or eb2 is not None and eb_counter == 1:
                                            eb_orientation = eb2
                                            eb_counter = 3
                                            if 's' in eb_orientation or 'S' in eb_orientation:
                                                S_COUNT.append(eb_orientation)
                                            if 'l' in eb_orientation or 'L' in eb_orientation:
                                                L_COUNT.append(eb_orientation)
                                            print(SKU_NUMBER + " :: " + eb_orientation)
                                        if eb3 is not None and eb_counter == 3:
                                            eb_orientation = eb3
                                            eb_counter = 4
                                            if 's' in eb_orientation or 'S' in eb_orientation:
                                                S_COUNT.append(eb_orientation)
                                            if 'l' in eb_orientation or 'L' in eb_orientation:
                                                L_COUNT.append(eb_orientation)
                                            print(SKU_NUMBER + " :: " + eb_orientation)
                                        if eb4 is not None and eb_counter == 4:
                                            eb_orientation = eb4
                                            eb_counter = 1
                                            if 's' in eb_orientation or 'S' in eb_orientation:
                                                S_COUNT.append(eb_orientation)
                                            if 'l' in eb_orientation or 'L' in eb_orientation:
                                                L_COUNT.append(eb_orientation)
                                            print(SKU_NUMBER + " :: " + eb_orientation)
                                    else:
                                        print(SKU_NUMBER)
                                    # Continue statement to skip over edgebanding for glass shelves
                                    if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                        continue
                                    # Continue statement to skip over edgebanding for Upgraded Panels doors
                                    if style_name != 'None' and sku_value[:2] in 'EB' and 'Slab Door' not in style_name:
                                        continue
                                    if style_name != 'None' and sku_value[:2] not in 'EB' and 'Slab Door' not in style_name and sku_value[:2] not in 'GL':
                                        # pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                        if is_glaze:
                                            if glaze_color != 'None' and glaze_style != 'None':
                                                if glass_color != 'None':
                                                    PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                                else:
                                                    PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                            else:
                                                if glass_color != 'None':
                                                    PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + ")"
                                                else:
                                                    PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + ")"
                                        else:
                                            if glass_color != 'None':
                                                PART_NAME = dcname + " (" + style_name + "(" + glass_color + "))"
                                            else:
                                                PART_NAME = dcname + " (" + style_name + ")"

                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                        if 'Slab Door' in style_name:
                                            MATERIAL_PARTS_LIST.append(
                                                [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2],
                                                 PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8],
                                                 eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled])
                                        else:
                                            UPGRADED_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                        #Added for user to price the Glass within the upgraded panel parts separately
                                        # if 'Glass' in PART_NAME or 'Glass' in style_name:
                                        #     # PART_NAME = style_name + " (" + glass_color + ")"
                                        #     PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                        #     if 'GL' not in sku_value:
                                        #         SKU_NUMBER = get_glass_sku(glass_color)
                                        #         pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                        #     GLASS_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                    else:
                                        if 'Glass' in PART_NAME or 'Glass' in style_name:
                                            if 'Glass Shelf' in PART_NAME:
                                                PART_NAME = dcname + " (" + glass_color + ")"
                                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name=None, is_glaze=False, glaze_style=None, glaze_color=None, part_name=PART_NAME)
                                                GLASS_PARTS_LIST.append(
                                                    [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME,
                                                    QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                                            else:
                                                print("    Glass SKU: " + sku_value, "Glass Color: " + glass_color, "Style Name: " + style_name, dcname)
                                                PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                                GLASS_THICKNESS = get_glass_thickness_by_sku(sku_value)
                                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail,upgrade_color)
                                                GLASS_PARTS_LIST.append(
                                                    [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME,
                                                    QUANTITY, LENGTH, WIDTH, GLASS_THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1],
                                                    center_rail, style_name])
                                        else:
                                            if upgrade_color is not None:
                                                PART_NAME = dcname + " (" + upgrade_color + ")"
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation, center_rail, upgrade_color)
                                            MATERIAL_PARTS_LIST.append(
                                                [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2],
                                                 PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8],
                                                 eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled])
                                if sku_value[:2] in hardware_types:
                                    SKU_NUMBER = sku_value
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                    PART_NAME = pricing_info[2]
                                    HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                                if sku_value[:2] in accessory_types:
                                    SKU_NUMBER = sku_value
                                    if 'rod'.upper() in PART_NAME.upper():
                                        LENGTH = rod_length
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                    else:
                                        LENGTH = lenx
                                        WIDTH = leny
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                    if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                        PART_NAME = pricing_info[2][:-10]
                                    else:
                                        PART_NAME = pricing_info[2]
                                    ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, WIDTH, QUANTITY, pricing_info[0], pricing_info[1], pricing_info[8]])
                                if sku_value[:2] in special_order_types:
                                    LENGTH = lenx
                                    WIDTH = leny
                                    if 'Glass' in PART_NAME or 'Glass' in style_name:
                                        PART_NAME = dcname + " (" + glass_color + ")"
                                    SKU_NUMBER = sku_value
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, None, center_rail)
                                    SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                    
                        EDGEBANDING.append([PART_LABEL_ID, len(S_COUNT), len(L_COUNT)])
                        S_COUNT.clear()
                        L_COUNT.clear()
                eb_counter = 1
            LENGTH = 0
            WIDTH = 0
            THICKNESS = 0

            for assembly in assembly.findall("Assembly"):

                for part in assembly.findall("Part"):
                    DESCRIPTION = description.text
                    PART_LABEL_ID = part.attrib.get('LabelID')

                    for quantity in part.findall("Quantity"):
                        QUANTITY = quantity.text

                    for part_name in part.findall("Name"):
                        if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                            PART_NAME = "Cleat"
                        elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                            PART_NAME = "Drawer Side"
                        elif part_name.text == "Left Door" or part_name.text == "Right Door":
                            PART_NAME = "Door"
                        else:
                            PART_NAME = part_name.text

                        if part_name.text is not None:
                            for width in part.findall("Width"):
                                WIDTH = width.text

                            for length in part.findall("Length"):
                                LENGTH = length.text

                            for thickness in part.findall("Thickness"):
                                THICKNESS = thickness.text

                            for part_type in part.findall("Type"):
                                PART_TYPE = part_type.text

                    for label in mfg.findall("Label"):
                        label_id = label.attrib.get("ID", None)
                        # part_id = part.attrib.get("PartID", None)
                        dcname = None
                        eb1 = None
                        eb2 = None
                        eb3 = None
                        eb4 = None
                        rod_length = None
                        style_name = 'None'
                        glaze_style = 'None'
                        glaze_color = 'None'
                        is_glaze = False
                        glass_color = 'None'
                        wall_name = None
                        center_rail = 'None'
                        lenx = 0
                        leny = 0
                        upgrade_color = None
                        drill_pattern = "-"
                        pull_dim = "-"
                        BoxAssembled = 'False'
                        if PART_LABEL_ID == label_id:
                            print(PART_NAME + " :: " + PART_LABEL_ID)
                            test_iterator_for_assembled = iter(label)
                            for _ in range(int(len(label) / 3)):
                                sku_value = None
                                NAME = next(test_iterator_for_assembled).text
                                TYPE = next(test_iterator_for_assembled).text
                                VALUE = next(test_iterator_for_assembled).text
                                if NAME == 'BoxAssembled':
                                    BoxAssembled = VALUE
                            # Use the iterator to deal with repeated children and still in order
                            label_iterator = iter(label)
                            for _ in range(int(len(label) / 3)):
                                sku_value = None
                                NAME = next(label_iterator).text
                                TYPE = next(label_iterator).text
                                VALUE = next(label_iterator).text
                                if NAME == 'dcname':
                                    if VALUE == "Top Cleat" or VALUE == "Bottom Cleat":
                                        VALUE = "Cleat"
                                    if VALUE == "Left Drawer Side" or VALUE == "Right Drawer Side":
                                        VALUE = "Drawer Side"
                                    if VALUE == "Left Door" or VALUE == "Right Door":
                                        VALUE = "Door"
                                    dcname = VALUE
                                if 'sku' in NAME:
                                    sku_value = VALUE
                                    if sku_value == 'Unknown':
                                        sku_value = 'SO-0000001'
                                    if NAME == 'sku':
                                        if sku_value == 'None':
                                            sku_value = 'SO-0000001'
                                if NAME == 'wallname':
                                    wall_name = VALUE
                                if NAME == 'lenx':
                                    lenx = VALUE
                                if NAME == 'leny':
                                    leny = VALUE
                                if NAME == 'upgradeoption':
                                    upgrade_color = VALUE
                                if NAME == 'edgeband1':
                                    eb1 = VALUE
                                if NAME == 'edgeband2':
                                    eb2 = VALUE
                                if NAME == 'edgeband3':
                                    eb3 = VALUE
                                if NAME == 'edgeband4':
                                    eb4 = VALUE
                                if 'rod'.upper() in PART_NAME.upper():
                                    if NAME == 'lenx':
                                        rod_length = VALUE
                                if NAME == 'style':
                                    style_name = VALUE
                                if NAME == 'glazecolor':
                                    glaze_color = VALUE
                                    if glaze_color != 'None':
                                        is_glaze = True
                                if NAME == 'glazestyle':
                                    glaze_style = VALUE
                                    if glaze_style != 'None':
                                        is_glaze = True
                                if NAME == 'glasscolor':
                                    glass_color = VALUE
                                if NAME == 'has_center_rail':
                                    center_rail = VALUE
                                if NAME == 'drill_pattern':
                                    drill_pattern = VALUE
                                if NAME == 'pull_dim':
                                    pull_dim = VALUE
                                if NAME == 'BoxAssembled':
                                    BoxAssembled = VALUE
                            # for sku_value in label.findall("Value"):
                                if sku_value is not None:
                                    if sku_value[:2] in material_types:
                                        SKU_NUMBER = sku_value
                                        if sku_value[:2] in 'EB':
                                            if eb1 is not None and eb_counter == 1:
                                                eb_orientation = eb1
                                                eb_counter = 2
                                                if 's' in eb_orientation or 'S' in eb_orientation:
                                                    S_COUNT.append(eb_orientation)
                                                if 'l' in eb_orientation or 'L' in eb_orientation:
                                                    L_COUNT.append(eb_orientation)
                                                print(SKU_NUMBER + " :: " + eb_orientation)
                                            if eb2 is not None and eb_counter == 2 or eb2 is not None and eb_counter == 1:
                                                eb_orientation = eb2
                                                eb_counter = 3
                                                if 's' in eb_orientation or 'S' in eb_orientation:
                                                    S_COUNT.append(eb_orientation)
                                                if 'l' in eb_orientation or 'L' in eb_orientation:
                                                    L_COUNT.append(eb_orientation)
                                                print(SKU_NUMBER + " :: " + eb_orientation)
                                            if eb3 is not None and eb_counter == 3:
                                                eb_orientation = eb3
                                                eb_counter = 4
                                                if 's' in eb_orientation or 'S' in eb_orientation:
                                                    S_COUNT.append(eb_orientation)
                                                if 'l' in eb_orientation or 'L' in eb_orientation:
                                                    L_COUNT.append(eb_orientation)
                                                print(SKU_NUMBER + " :: " + eb_orientation)
                                            if eb4 is not None and eb_counter == 4:
                                                eb_orientation = eb4
                                                eb_counter = 1
                                                if 's' in eb_orientation or 'S' in eb_orientation:
                                                    S_COUNT.append(eb_orientation)
                                                if 'l' in eb_orientation or 'L' in eb_orientation:
                                                    L_COUNT.append(eb_orientation)
                                                print(SKU_NUMBER + " :: " + eb_orientation)
                                        else:
                                            print(SKU_NUMBER)
                                        # Continue statement to skip over edgebanding for glass shelves
                                        if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                            continue
                                        # Continue statement to skip over edgebanding for Upgraded Panels doors
                                        if style_name != 'None' and sku_value[:2] in 'EB' and 'Slab Door' not in style_name:
                                            continue
                                        if style_name != 'None' and sku_value[:2] not in 'EB' and 'Slab Door' not in style_name and sku_value[:2] not in 'GL':
                                            # pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                            if is_glaze:
                                                if glaze_color != 'None' and glaze_style != 'None':
                                                    if glass_color != 'None':
                                                        PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                                    else:
                                                        PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                                else:
                                                    if glass_color != 'None':
                                                        PART_NAME = dcname + " (" + style_name + "(" + glass_color + ")" + ", Glazed " + glaze_color + ")"
                                                    else:
                                                        PART_NAME = dcname + " (" + style_name + ", Glazed " + glaze_color + ")"
                                            else:
                                                if glass_color != 'None':
                                                    PART_NAME = dcname + " (" + style_name + "(" + glass_color + "))"
                                                else:
                                                    PART_NAME = dcname + " (" + style_name + ")"

                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                            if 'Slab Door' in style_name:
                                                MATERIAL_PARTS_LIST.append(
                                                    [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2],
                                                     PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8],
                                                     eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled])
                                            else:
                                                UPGRADED_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                            #Added for user to price the Glass within the upgraded panel parts separately
                                            # if 'Glass' in PART_NAME or 'Glass' in style_name:
                                            #     # PART_NAME = style_name + " (" + glass_color + ")"
                                            #     PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                            #     if 'GL' not in sku_value:
                                            #         SKU_NUMBER = get_glass_sku(glass_color)
                                            #         pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail)
                                            #     GLASS_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                                        else:
                                            if 'Glass' in PART_NAME or 'Glass' in style_name:
                                                if 'Glass Shelf' in PART_NAME:
                                                    PART_NAME = dcname + " (" + glass_color + ")"
                                                else:
                                                    print("    Glass SKU: " + sku_value, "Glass Color: " + glass_color, "Style Name: " + style_name, dcname)
                                                    PART_NAME = "Inset Glass" + " (" + glass_color + ")"
                                                    GLASS_THICKNESS = get_glass_thickness_by_sku(sku_value)
                                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME, None, center_rail,upgrade_color)
                                                # pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation, center_rail, upgrade_color)
                                                GLASS_PARTS_LIST.append(
                                                    [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME,
                                                     QUANTITY, LENGTH, WIDTH, GLASS_THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1],
                                                     center_rail, style_name])
                                            else:
                                                if upgrade_color is not None:
                                                    PART_NAME = dcname + " (" + upgrade_color + ")"
                                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation, center_rail, upgrade_color, BoxAssembled)
                                                MATERIAL_PARTS_LIST.append(
                                                    [DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2],
                                                     PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8],
                                                     eb_orientation, pricing_info[7], pricing_info[1], wall_name, drill_pattern, pull_dim, BoxAssembled])
                                    if sku_value[:2] in hardware_types:
                                        SKU_NUMBER = sku_value
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                        PART_NAME = pricing_info[2]
                                        HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                                    if sku_value[:2] in accessory_types:
                                        SKU_NUMBER = sku_value
                                        if 'rod'.upper() in PART_NAME.upper():
                                            LENGTH = rod_length
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                        else:
                                            LENGTH = lenx
                                            WIDTH = leny
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                        if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                            PART_NAME = pricing_info[2][:-10]
                                        else:
                                            PART_NAME = pricing_info[2]
                                        ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, WIDTH, QUANTITY, pricing_info[0], pricing_info[1], pricing_info[8]])
                                    if sku_value[:2] in special_order_types:
                                        LENGTH = lenx
                                        WIDTH = leny
                                        if 'Glass' in PART_NAME or 'Glass' in style_name:
                                            PART_NAME = dcname + " (" + glass_color + ")"
                                        SKU_NUMBER = sku_value
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, None, center_rail)
                                        SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1], center_rail])
                    
                            EDGEBANDING.append([PART_LABEL_ID, len(S_COUNT), len(L_COUNT)])
                            S_COUNT.clear()
                            L_COUNT.clear()
                    eb_counter = 1
                LENGTH = 0
                WIDTH = 0
                THICKNESS = 0

        R_ROOM_TOTAL_PRICE = sum(map(float, R_MATERIAL_PRICES)) + sum(map(float, R_HARDWARE_PRICES)) + sum(map(float, R_ACCESSORY_PRICES)) + sum(map(float, R_UPGRADED_PANEL_PRICES)) + sum(map(float, R_GLASS_PRICES))
        R_ROOM_PRICING_LIST.append([
            DESCRIPTION,
            sum(map(float, R_MATERIAL_SQUARE_FOOTAGE)),
            sum(map(float, R_MATERIAL_LINEAR_FOOTAGE)),
            sum(map(float, R_MATERIAL_PRICES)),
            sum(map(float, R_HARDWARE_PRICES)),
            sum(map(float, R_ACCESSORY_PRICES)),
            sum(map(float, R_UPGRADED_PANEL_PRICES)),
            len(R_SPECIAL_ORDER_PRICES),
            sum(map(float, R_LABOR_PRICES)),
            sum(map(float, R_GLASS_PRICES)),
            R_ROOM_TOTAL_PRICE,
            ROOM_DRAWING_NUMBER,
            ])

        R_PROJECT_TOTAL_HARDWARE.append(sum(map(float, R_HARDWARE_PRICES)))
        R_PROJECT_TOTAL_ACCESSORIES.append(sum(map(float, R_ACCESSORY_PRICES)))
        R_PROJECT_TOTAL_SQUARE_FOOTAGE.append(sum(map(float, R_MATERIAL_SQUARE_FOOTAGE)))
        R_PROJECT_TOTAL_LINEAR_FOOTAGE.append(sum(map(float, R_MATERIAL_LINEAR_FOOTAGE)))
        R_PROJECT_TOTAL_MATERIAL.append(sum(map(float, R_MATERIAL_PRICES)))
        R_PROJECT_TOTAL_UPGRADED_PANEL.append(sum(map(float, R_UPGRADED_PANEL_PRICES)))
        R_PROJECT_TOTAL_GLASS.append(sum(map(float, R_GLASS_PRICES)))
        R_PROJECT_TOTAL_LABOR.append(sum(map(float, R_LABOR_PRICES)))
        R_PROJECT_TOTAL_PRICE.append(R_ROOM_TOTAL_PRICE)
        R_ROOM_TOTAL_PRICE = 0
        R_MATERIAL_PRICES.clear()
        R_HARDWARE_PRICES.clear()
        R_ACCESSORY_PRICES.clear()
        R_UPGRADED_PANEL_PRICES.clear()
        R_SPECIAL_ORDER_PRICES.clear()
        R_LABOR_PRICES.clear()
        R_GLASS_PRICES.clear()
        R_MATERIAL_SQUARE_FOOTAGE.clear()
        R_MATERIAL_LINEAR_FOOTAGE.clear()

        F_ROOM_TOTAL_PRICE = sum(map(float, F_MATERIAL_PRICES)) + sum(map(float, F_HARDWARE_PRICES)) + sum(map(float, F_ACCESSORY_PRICES)) + sum(map(float, F_UPGRADED_PANEL_PRICES)) + sum(map(float, F_GLASS_PRICES))
        F_ROOM_PRICING_LIST.append([DESCRIPTION, sum(map(float, F_MATERIAL_SQUARE_FOOTAGE)), sum(map(float, F_MATERIAL_LINEAR_FOOTAGE)), sum(map(float, F_MATERIAL_PRICES)), sum(map(float, F_HARDWARE_PRICES)), sum(map(float, F_ACCESSORY_PRICES)), sum(map(float, F_UPGRADED_PANEL_PRICES)), len(F_SPECIAL_ORDER_PRICES), sum(map(float, F_LABOR_PRICES)), sum(map(float, F_GLASS_PRICES)), F_ROOM_TOTAL_PRICE])
        F_PROJECT_TOTAL_HARDWARE.append(sum(map(float, F_HARDWARE_PRICES)))
        F_PROJECT_TOTAL_ACCESSORIES.append(sum(map(float, F_ACCESSORY_PRICES)))
        F_PROJECT_TOTAL_SQUARE_FOOTAGE.append(sum(map(float, F_MATERIAL_SQUARE_FOOTAGE)))
        F_PROJECT_TOTAL_LINEAR_FOOTAGE.append(sum(map(float, F_MATERIAL_LINEAR_FOOTAGE)))
        F_PROJECT_TOTAL_MATERIAL.append(sum(map(float, F_MATERIAL_PRICES)))
        F_PROJECT_TOTAL_UPGRADED_PANEL.append(sum(map(float, F_UPGRADED_PANEL_PRICES)))
        F_PROJECT_TOTAL_GLASS.append(sum(map(float, F_GLASS_PRICES)))
        F_PROJECT_TOTAL_LABOR.append(sum(map(float, F_LABOR_PRICES)))
        F_PROJECT_TOTAL_PRICE.append(F_ROOM_TOTAL_PRICE)
        F_ROOM_TOTAL_PRICE = 0
        F_MATERIAL_PRICES.clear()
        F_HARDWARE_PRICES.clear()
        F_ACCESSORY_PRICES.clear()
        F_UPGRADED_PANEL_PRICES.clear()
        F_SPECIAL_ORDER_PRICES.clear()
        F_LABOR_PRICES.clear()
        F_GLASS_PRICES.clear()
        F_MATERIAL_SQUARE_FOOTAGE.clear()
        F_MATERIAL_LINEAR_FOOTAGE.clear()

    print("Calculate Price COMPLETE")
    if bpy.context.window:
        bpy.context.window.cursor_set("DEFAULT")

    if COS_FLAG:
        generate_retail_parts_list()
        generate_franchise_parts_list()

    return xml_file_exists


def calc_room_total_sq_ft(series):
    return numpy.sum(series['SQUARE_FT'] * series['QUANTITY'])


class SNAP_OT_Calculate_Price(Operator):
    bl_idname = PRICING_PROPERTY_NAMESPACE + ".calculate_price"
    bl_label = "Calculate Price"
    bl_description = "Calculate Price for Project"

    tmp_filename = "export_temp.py"
    xml_filename = "snap_job.xml"
    proj_dir: StringProperty(name="Project Directory", subtype='DIR_PATH')

    def invoke(self, context, event):
        wm = context.window_manager
        return wm.invoke_props_dialog(self, width=300)

    def draw(self, context):
        layout = self.layout
        props = context.window_manager.sn_project
        proj = props.get_project()
        layout.label(text="Project: {}".format(proj.name))
        box = layout.box()

        for room in proj.rooms:
            if len(proj.rooms) == 1:
                room.selected = True
            col = box.column(align=True)
            row = col.row()
            row.prop(room, "selected", text="")
            row.label(text=room.name)

        row = layout.row()
        row.operator(
            "sn_project_pricing.select_all_rooms", text="Select All", icon='CHECKBOX_HLT').select_all = True
        row.operator(
            "sn_project_pricing.select_all_rooms", text="Deselect All", icon='CHECKBOX_DEHLT').select_all = False

    def create_prep_script(self):
        nrm_dir = self.proj_dir.replace("\\", "/")
        file = open(os.path.join(bpy.app.tempdir, self.tmp_filename), 'w')
        file.write("import bpy\n")
        file.write("bpy.ops.sn_export.export_xml('INVOKE_DEFAULT', xml_path='{}')\n".format(nrm_dir))
        file.close()
        return os.path.join(bpy.app.tempdir, self.tmp_filename)
        
    def execute(self, context):
        bpy.ops.wm.save_mainfile()
        if context.window:
            bpy.context.window.cursor_set("WAIT")
            debug_mode = context.preferences.addons["snap"].preferences.debug_mode
            debug_mac = context.preferences.addons["snap"].preferences.debug_mac
            proj_props = bpy.context.window_manager.sn_project
            proj_name = proj_props.projects[proj_props.project_index].name
            path = os.path.join(sn_xml.get_project_dir(), proj_name, self.xml_filename)
            proj = proj_props.projects[proj_props.project_index]

        if os.path.exists(path):
            os.remove(path)

        self.proj_dir = os.path.join(sn_xml.get_project_dir(), proj_name)
        script_path = self.create_prep_script()

        # Call blender in background and run XML export on each room file in project
        for room in proj.rooms:
            if room.selected:
                subprocess.call(bpy.app.binary_path + ' "' + room.file_path + '" -b --python "' + script_path + '"')

        if debug_mode and debug_mac:
            bpy.ops.wm.open_mainfile(filepath=bpy.data.filepath)
            if "Machining" in bpy.data.collections:
                for obj in bpy.data.collections["Machining"].objects:
                    obj.display_type = 'WIRE'

        props = get_pricing_props()
        props.reset()
        props.calculate_price(context)
        return {'FINISHED'}


class SNAP_OT_Select_All_Rooms_Pricing(Operator):
    bl_idname = "sn_project_pricing.select_all_rooms"
    bl_label = "Select All Rooms"
    bl_description = "This will select all of the rooms in the project"

    select_all: BoolProperty(name="Select All", default=True)

    @classmethod
    def poll(cls, context):
        return True

    def execute(self, context):
        props = context.window_manager.sn_project
        proj = props.projects[props.project_index]

        for room in proj.rooms:
            room.selected = self.select_all

        return{'FINISHED'}


class SNAP_OT_Price_COS_XML(Operator):
    bl_idname = "sn_project_pricing.price_cos_xml"
    bl_label = "Price COS XML"
    bl_description = "This will process and price a COS generated XML file"

    filename: StringProperty(name="Project File Name", description="Project file name to import")
    filepath: StringProperty(name="Project Path", description="Project path to import", subtype="FILE_PATH")
    directory: StringProperty(name="Project File Directory Name", description="Project file directory name")
    # ImportHelper mixin class uses this
    filename_ext = ".blend"
    filter_glob: StringProperty(default="*.xml", options={'HIDDEN'}, maxlen=255)

    @classmethod
    def poll(cls, context):
        return True

    def invoke(self, context, event):
        context.window_manager.fileselect_add(self)
        return {'RUNNING_MODAL'}

    def execute(self, context):
        props = get_pricing_props()
        props.reset()
        blender_path = os.path.dirname(bpy.app.binary_path)
        file_exists = False

        if pathlib.Path(self.filename).suffix == ".xml":
            file_exists = calculate_project_price(self.filepath, cos_flag=True)

        if not file_exists:
            message = "This is not a valid XML file!: {}".format(self.filename)
            bpy.ops.snap.message_box('INVOKE_DEFAULT', message=message, icon='ERROR')

        return {'FINISHED'}


class SNAP_OT_Test_COS_XML(Operator):
    bl_idname = "sn_project_pricing.test_cos_xml"
    bl_label = "Test COS XML"
    bl_description = "This will process and price the test COS XML file"


    @classmethod
    def poll(cls, context):
        return True

    def execute(self, context):
        props = get_pricing_props()
        props.reset()
        blender_path = os.path.dirname(bpy.app.binary_path)
        test_xml_path = os.path.join(blender_path,"3.0\\scripts\\addons\\snap\\tests\\pricing\\cos_test.xml")
        file_exists = False

        if pathlib.Path(test_xml_path).suffix == ".xml":
            file_exists = calculate_project_price(test_xml_path, cos_flag=True)

        if not file_exists:
            message = "This is not a valid XML file!: {}".format(test_xml_path)
            bpy.ops.snap.message_box('INVOKE_DEFAULT', message=message, icon='ERROR')

        return {'FINISHED'}


class SNAP_PROPS_Pricing(bpy.types.PropertyGroup):

    pricing_tabs: EnumProperty(
        name="Pricing Tabs",
        items=[('RETAIL', "Retail Pricing", 'View Pricing Information'), 
               ('FRANCHISE', "Franchise Pricing", 'View Pricing information')])

    export_pricing_parts_list: BoolProperty(
        name="Export Pricing Parts List (.xlsx)",
        description="Export a list of all parts being priced",
        default=True)

    use_tearout_pricing: BoolProperty(
        name="Include Tearout Pricing",
        description="Select Tearout Pricing Options",
        default=False)
    customer_tearout: BoolProperty(
        name="Customer Tearout",
        description="Customer will perform tearout. No additional charge.",
        default=False)
    light_same_tearout: BoolProperty(
        name="Light Same-Day",
        description="No additional charge under 25 LF.",
        default=False)
    heavy_same_bool_0: BoolProperty(
        name="Heavy Same-Day",
        description="Heavy Same-Day Tearout. $5 per LF",
        default=False)
    heavy_same_int_0: IntProperty(
        name="Heavy Same-Day Tearout", 
        description="Heavy Same-Day Tearout. $5 per LF")
    heavy_same_bool_1: BoolProperty(
        name="Extra Heavy Same-Day",
        description="Extra Heavy Same-Day Tearout. $10 per LF",
        default=False)
    heavy_same_int_1: IntProperty(
        name="Extra Heavy Same-Day Tearout", 
        description="Extra Heavy Same-Day Tearout. $10 per LF")
    light_prior_tearout: BoolProperty(
        name="Light Prior Day",
        description="$75 trip charge",
        default=False)
    heavy_prior_bool_0: BoolProperty(
        name="Heavy Prior-Day",
        description="Heavy Prior-Day Tearout. $5 per LF",
        default=False)
    heavy_prior_int_0: IntProperty(
        name="Heavy Prior Day Tearout", 
        description="Heavy Prior Day Tearout. $5 per LF + $75 trip charge")
    heavy_prior_bool_1: BoolProperty(
        name="Extra Heavy Prior-Day",
        description="Extra Heavy Prior-Day Tearout. $10 per LF + $75 trip charge",
        default=False)
    heavy_prior_int_1: IntProperty(
        name="Extra Heavy Prior Day Tearout", 
        description="Extra Heavy Prior Day Tearout. $10 per LF + $75 trip charge")
    
    
    def reset(self):
        '''
            This function resets all of the pricing properties back to their default
            This should be run before you calculate the price to ensure there is no
            values stored from a previous calculation
        '''
        R_MATERIAL_PRICES.clear()
        R_HARDWARE_PRICES.clear()
        R_ACCESSORY_PRICES.clear()
        R_UPGRADED_PANEL_PRICES.clear()
        R_SPECIAL_ORDER_PRICES.clear()
        R_LABOR_PRICES.clear()
        R_GLASS_PRICES.clear()
        R_MATERIAL_SQUARE_FOOTAGE.clear()
        R_MATERIAL_LINEAR_FOOTAGE.clear()
        R_PROJECT_TOTAL_HARDWARE.clear()
        R_PROJECT_TOTAL_ACCESSORIES.clear()
        R_PROJECT_TOTAL_MATERIAL.clear()
        R_PROJECT_TOTAL_UPGRADED_PANEL.clear()
        R_PROJECT_TOTAL_SQUARE_FOOTAGE.clear()
        R_PROJECT_TOTAL_LINEAR_FOOTAGE.clear()
        R_PROJECT_TOTAL_PRICE.clear()
        R_PROJECT_TOTAL_GLASS.clear()
        R_ROOM_PRICING_LIST.clear()

        MATERIAL_PARTS_LIST.clear()
        ACCESSORY_PARTS_LIST.clear()
        HARDWARE_PARTS_LIST.clear()
        UPGRADED_PANEL_PARTS_LIST.clear()
        SPECIAL_ORDER_PARTS_LIST.clear()
        GLASS_PARTS_LIST.clear()

        F_MATERIAL_PRICES.clear()
        F_HARDWARE_PRICES.clear()
        F_ACCESSORY_PRICES.clear()
        F_UPGRADED_PANEL_PRICES.clear()
        F_SPECIAL_ORDER_PRICES.clear()
        F_LABOR_PRICES.clear()
        F_GLASS_PRICES.clear()
        F_MATERIAL_SQUARE_FOOTAGE.clear()
        F_MATERIAL_LINEAR_FOOTAGE.clear()
        F_PROJECT_TOTAL_HARDWARE.clear()
        F_PROJECT_TOTAL_ACCESSORIES.clear()
        F_PROJECT_TOTAL_MATERIAL.clear()
        F_PROJECT_TOTAL_UPGRADED_PANEL.clear()
        F_PROJECT_TOTAL_SQUARE_FOOTAGE.clear()
        F_PROJECT_TOTAL_LINEAR_FOOTAGE.clear()
        F_PROJECT_TOTAL_PRICE.clear()
        F_PROJECT_TOTAL_GLASS.clear()
        F_ROOM_PRICING_LIST.clear()

    def calculate_price(self, context):
        '''
            This function calculates the price of the project from the xml export
        '''
        xml_file = get_project_xml(self)
        if xml_file is not None:
            calculate_project_price(xml_file)
            props = get_pricing_props()
            snap_prefs = bpy.context.preferences.addons['snap'].preferences
            if props.export_pricing_parts_list:
                generate_retail_parts_list()

                if snap_prefs.enable_franchise_pricing or snap_prefs.debug_mode:
                    generate_franchise_parts_list()
        else:
            raise NameError("The 'snap_job.xml' was not created and returned a NoneType. Pricing NOT Generated.")

    def draw(self, layout):
        box = layout.box()
        row = box.row(align=True)
        row.operator(PRICING_PROPERTY_NAMESPACE + ".calculate_price",icon='FILE_TICK')

        admin_fee = get_admin_fee()

        if bpy.context.preferences.addons['snap'].preferences.debug_mode:
            row.menu('SNAP_MT_Pricing_Tools', text="", icon='DOWNARROW_HLT')

        row = layout.row(align=True)
        row.prop(self,'export_pricing_parts_list',text="Export Pricing Parts List (.xlsx)")

        # split = row.split()
        # row = split.row()
        # row = layout.row(align=True)
        # row.prop(self, "use_tearout_pricing", text="Include Tearout Pricing")
        # if self.use_tearout_pricing:
        #     box = layout.box()
        #     box.enabled = self.use_tearout_pricing
        #     row = box.row(align=True)
        #     row.prop(self, "customer_tearout")
        #     if self.customer_tearout:
        #         row.label(text="No Charge")
        #     row = box.row(align=True)
        #     row.prop(self, "light_same_tearout")
        #     if self.light_same_tearout:
        #         row.label(text="No Charge under 25 LF")
        #     row = box.row(align=True)
        #     row.prop(self, "heavy_same_bool_0")
        #     if self.heavy_same_bool_0:
        #         row.prop(self, "heavy_same_int_0",text="$")
        #     row = box.row(align=True)
        #     row.prop(self, "heavy_same_bool_1")
        #     if self.heavy_same_bool_1:
        #         row.prop(self, "heavy_same_int_1",text="$")
        #     row = box.row(align=True)
        #     row.prop(self, "light_prior_tearout")
        #     if self.light_prior_tearout:
        #         row.label(text="$75 trip charge")
        #     row = box.row(align=True)
        #     row.prop(self, "heavy_prior_bool_0")
        #     if self.heavy_prior_bool_0:
        #         row.prop(self, "heavy_prior_int_0",text="$")
        #     row = box.row(align=True)
        #     row.prop(self, "heavy_prior_bool_1")
        #     if self.heavy_prior_bool_1:
        #         row.prop(self, "heavy_prior_int_1",text="$")

        box = layout.box()
        main_col = box.column(align=True)
        row = main_col.row(align=True)
        row.scale_y = 1
        row.prop_enum(self, "pricing_tabs", 'RETAIL', icon='PREFERENCES', text="Retail Pricing")

        if self.pricing_tabs == 'RETAIL':
            box = main_col.box()
            col = box.column()
            for i in range(len(R_ROOM_PRICING_LIST)):
                col.label(text="Room Name: " + R_ROOM_PRICING_LIST[i][0])
                col.label(text="Special Order Items Count: " + str(R_ROOM_PRICING_LIST[i][7]),icon='BLANK1')
                col.label(text="Materials Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][3])),icon='BLANK1')
                col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][4])),icon='BLANK1')
                col.label(text="Acccessories Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][5])),icon='BLANK1')
                col.label(text="Upgraded Panels Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][6])),icon='BLANK1')
                col.label(text="Glass Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][9])),icon='BLANK1')
                # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][8]),icon='BLANK1')
                col.label(text="Room Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][10])),icon='BLANK1')
                col.label(text="Admin Fee: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][10] * admin_fee)),icon='BLANK1')
                col.label(text="Adjusted Room Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(R_ROOM_PRICING_LIST[i][10] + (R_ROOM_PRICING_LIST[i][10] * admin_fee))),icon='BLANK1')
                col.separator()

            col.separator()
            col.label(text="Project Totals: ")
            col.label(text="Room Count: " + str(len(R_ROOM_PRICING_LIST)),icon='BLANK1')
            col.label(text="Special Order Items Count: " + str(len(SPECIAL_ORDER_PARTS_LIST)),icon='BLANK1')
            col.label(text="Materials Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_MATERIAL)))),icon='BLANK1')
            col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_HARDWARE)))),icon='BLANK1')
            col.label(text="Accessories Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_ACCESSORIES)))),icon='BLANK1')
            col.label(text="Upgraded Panels Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_UPGRADED_PANEL)))),icon='BLANK1')
            col.label(text="Glass Price: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_GLASS)))),icon='BLANK1')
            # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_LABOR))),icon='BLANK1')
            col.label(text="Project Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_PRICE)))),icon='BLANK1')
            col.label(text="Admin Fee: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_PRICE)) * admin_fee)),icon='BLANK1')
            col.label(text="Adjusted Project Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, R_PROJECT_TOTAL_PRICE)) + (sum(map(float, R_PROJECT_TOTAL_PRICE)) * admin_fee))),icon='BLANK1')
            
            if len(SPECIAL_ORDER_PARTS_LIST) != 0:
                col.separator()
                col.label(text="This project contains special order items,")
                col.label(text="which are not being calculated in this pricing summary")

        if bpy.context.preferences.addons['snap'].preferences.enable_franchise_pricing:
            row.prop_enum(self, "pricing_tabs", 'FRANCHISE', icon='PREFERENCES', text="Franchise Pricing")

            if self.pricing_tabs == 'FRANCHISE':
                box = main_col.box()
                col = box.column()
                for i in range(len(F_ROOM_PRICING_LIST)):
                    col.label(text="Room Name: " + F_ROOM_PRICING_LIST[i][0])
                    col.label(text="Special Order Items Count: " + str(F_ROOM_PRICING_LIST[i][7]),icon='BLANK1')
                    col.label(text="Materials Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][3]),icon='BLANK1')
                    col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][4]),icon='BLANK1')
                    col.label(text="Acccessories Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][5]),icon='BLANK1')
                    col.label(text="Upgraded Panels Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][6]),icon='BLANK1')
                    # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][8]),icon='BLANK1')
                    col.label(text="Glass Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][9]),icon='BLANK1')
                    col.label(text="Room Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(F_ROOM_PRICING_LIST[i][10])),icon='BLANK1')
                    col.separator()

                col.separator()
                col.label(text="Project Totals: ")
                col.label(text="Room Count: " + str(len(F_ROOM_PRICING_LIST)),icon='BLANK1')
                col.label(text="Special Order Items Count: " + str(len(SPECIAL_ORDER_PARTS_LIST)),icon='BLANK1')
                col.label(text="Materials Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_MATERIAL))),icon='BLANK1')
                col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_HARDWARE))),icon='BLANK1')
                col.label(text="Accessories Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_ACCESSORIES))),icon='BLANK1')
                col.label(text="Upgraded Panels Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_UPGRADED_PANEL))),icon='BLANK1')
                # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_LABOR))),icon='BLANK1')
                col.label(text="Glass Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_GLASS))),icon='BLANK1')
                col.label(text="Project Subtotal: " + sn_unit.draw_dollar_price(numpy.math.ceil(sum(map(float, F_PROJECT_TOTAL_PRICE)))),icon='BLANK1')

                if len(SPECIAL_ORDER_PARTS_LIST) != 0:
                    col.separator()
                    col.label(text="This project contains special order items,")
                    col.label(text="which are not being calculated in this pricing summary") 


class SNAP_MT_Pricing_Tools(bpy.types.Menu):
    bl_label = "Pricing Debug Tools"

    def draw(self, context):
        layout = self.layout

        layout.operator("sn_project_pricing.price_cos_xml",
                        text="Price COS XML",
                        icon='FILE_FOLDER')

        layout.operator("sn_project_pricing.test_cos_xml",
                        text="Test COS XML pricing",
                        icon='FILE_FOLDER')


class SNAP_PT_Project_Pricing_Setup(Panel):
    bl_label = "Project Pricing"
    bl_space_type = "VIEW_3D"
    bl_region_type = "TOOLS"
    bl_context = "objectmode"
    bl_options = {'DEFAULT_CLOSED'}
    
    def draw_header(self, context):
        layout = self.layout
        layout.label(text='',icon='LINE_DATA')

    def draw(self, context):
        props = get_pricing_props()
        props.draw(self.layout)


# REGISTER CLASSES
def register():
    bpy.utils.register_class(SNAP_OT_Calculate_Price)
    bpy.utils.register_class(SNAP_OT_Select_All_Rooms_Pricing)
    bpy.utils.register_class(SNAP_OT_Price_COS_XML)
    bpy.utils.register_class(SNAP_OT_Test_COS_XML)
    bpy.utils.register_class(SNAP_MT_Pricing_Tools)
    bpy.utils.register_class(SNAP_PROPS_Pricing)
    bpy.utils.register_class(SNAP_PT_Project_Pricing_Setup)
    exec("bpy.types.Scene." + PRICING_PROPERTY_NAMESPACE + "= PointerProperty(type = SNAP_PROPS_Pricing)")


def unregister():
    bpy.utils.unregister_class(SNAP_PT_Project_Pricing_Setup)
    bpy.utils.unregister_class(SNAP_PROPS_Pricing)
    bpy.utils.unregister_class(SNAP_MT_Pricing_Tools)
    bpy.utils.unregister_class(SNAP_OT_Price_COS_XML)
    bpy.utils.unregister_class(SNAP_OT_Test_COS_XML)
    bpy.utils.unregister_class(SNAP_OT_Calculate_Price)
    bpy.utils.unregister_class(SNAP_OT_Select_All_Rooms_Pricing)
    exec("del bpy.types.Scene." + PRICING_PROPERTY_NAMESPACE)


# AUTO CALL OPERATOR ON SAVE
# @bpy.app.handlers.persistent
# def calculate_pricing(scene=None):
#     props = get_pricing_props()
#     if props.auto_calculate_on_save:
#         exec("bpy.ops." + PRICING_PROPERTY_NAMESPACE + ".calculate_price()")
# bpy.app.handlers.save_pre.append(calculate_pricing)
